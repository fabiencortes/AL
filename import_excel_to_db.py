# import_excel_to_db.py
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from database import get_connection

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Planning 2025.xlsx")


def _normalize_time(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip().replace("H", "h").replace("h", ":").replace(" ", "")
    if not s:
        return ""
    if isinstance(val, (pd.Timestamp, datetime)):
        try:
            return val.strftime("%H:%M")
        except Exception:
            pass
    if s.isdigit():
        if len(s) <= 2:
            try:
                h = int(s); m = 0
            except ValueError:
                return s
        else:
            try:
                h = int(s[:-2]); m = int(s[-2:])
            except ValueError:
                return s
        if 0 <= h <= 23 and 0 <= m <= 59:
            return f"{h:02d}:{m:02d}"
        return s
    if ":" in s:
        p = s.split(":")
        if len(p) >= 2:
            try:
                h = int(p[0]); m = int(p[1])
            except ValueError:
                return s
            if 0 <= h <= 23 and 0 <= m <= 59:
                return f"{h:02d}:{m:02d}"
        return s
    return s


def _normalize_date(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    try:
        d = pd.to_datetime(v, dayfirst=True, errors="coerce")
        if pd.isna(d):
            return str(v)
        return d.strftime("%d/%m/%Y")
    except Exception:
        return str(v)


def _detect_groupage_partage_from_colors(df: pd.DataFrame) -> pd.DataFrame:
    """Déduit GROUPAGE / PARTAGE à partir des couleurs Excel de DATE/HEURE."""
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
    except Exception:
        return df

    if "Feuil1" in wb.sheetnames:
        ws = wb["Feuil1"]
    else:
        ws = wb.active

    headers = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.col_idx

    date_col = headers.get("DATE")
    heure_col = headers.get("HEURE")
    if not date_col or not heure_col:
        return df

    # codes jaune (RGB ou index Excel)
    YELLOW_RGB = {"FFFF00", "FFF9C4", "FFFFFF00"}
    YELLOW_INDEXES = {6, 36, 44}  # 6 = jaune standard Excel

    def _is_yellow(cell):
        try:
            fill = cell.fill
            if not fill:
                return False
            col = fill.fgColor

            # RGB direct
            if getattr(col, "rgb", None):
                rgb = col.rgb.upper().replace("#", "")
                if len(rgb) == 8:
                    rgb = rgb[2:]
                return rgb in YELLOW_RGB

            # Couleur indexée
            if getattr(col, "indexed", None) is not None:
                return col.indexed in YELLOW_INDEXES

            return False
        except Exception:
            return False

    n_rows_df = len(df)
    group_flags = ["0"] * n_rows_df
    partage_flags = ["0"] * n_rows_df

    # ATTENTION : df.dropna() a retiré des lignes → on aligne par index + 2
    # On suppose que l'ordre des lignes non vides correspond encore
    for i, idx in enumerate(df.index):
        excel_row = i + 2  # 2 car 1 = entête
        c_date = ws.cell(row=excel_row, column=date_col)
        c_heure = ws.cell(row=excel_row, column=heure_col)
        is_date_yellow = _is_yellow(c_date)
        is_heure_yellow = _is_yellow(c_heure)

        if is_date_yellow and is_heure_yellow:
            group_flags[i] = "1"
        elif (not is_date_yellow) and is_heure_yellow:
            partage_flags[i] = "1"

    # S'assurer que les colonnes existent
    if "GROUPAGE" not in df.columns:
        df["GROUPAGE"] = "0"
    if "PARTAGE" not in df.columns:
        df["PARTAGE"] = "0"

    for i, idx in enumerate(df.index):
        if group_flags[i] == "1":
            df.at[idx, "GROUPAGE"] = "1"
        if partage_flags[i] == "1":
            df.at[idx, "PARTAGE"] = "1"

    return df


def import_planning_from_feuil1():
    """Recharge complètement la table planning depuis Feuil1."""
    try:
        df = pd.read_excel(
            EXCEL_FILE,
            sheet_name="Feuil1",
            engine="openpyxl",
        )
    except Exception as e:
        raise RuntimeError(f"Erreur lecture Excel ({EXCEL_FILE}) : {e}")

    df = df.dropna(how="all")
    if df.empty:
        raise RuntimeError("Feuil1 est vide après nettoyage.")

    df.columns = [str(c).strip() for c in df.columns]

    if "DATE" not in df.columns:
        raise RuntimeError("Colonne 'DATE' manquante dans Feuil1.")

    df["DATE"] = df["DATE"].apply(_normalize_date)

    if "HEURE" in df.columns:
        df["HEURE"] = df["HEURE"].apply(_normalize_time)
    if "²²²²" in df.columns:
        df["²²²²"] = df["²²²²"].apply(_normalize_time)

    if "CH" not in df.columns:
        df["CH"] = ""

    # IMPORTANT : créer les colonnes GROUPAGE / PARTAGE si absentes
    if "GROUPAGE" not in df.columns:
        df["GROUPAGE"] = "0"
    if "PARTAGE" not in df.columns:
        df["PARTAGE"] = "0"

    # puis les remplir depuis les couleurs
    df = _detect_groupage_partage_from_colors(df)

    df = df.fillna("")

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("DROP TABLE IF EXISTS planning")

        col_defs = ", ".join(f'"{c}" TEXT' for c in df.columns)
        create_sql = f"""
        CREATE TABLE planning (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {col_defs}
        );
        """
        cur.execute(create_sql)

        col_list_sql = ", ".join(f'"{c}"' for c in df.columns)
        placeholders = ", ".join("?" for _ in df.columns)
        insert_sql = f"INSERT INTO planning ({col_list_sql}) VALUES ({placeholders})"

        for _, row in df.iterrows():
            values = [str(row[c]) if row[c] != "" else "" for c in df.columns]
            cur.execute(insert_sql, values)

        conn.commit()
