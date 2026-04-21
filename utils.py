import pandas as pd
import hashlib
from io import BytesIO
from openpyxl import load_workbook
import streamlit as st





# ======================================================
# 🐞 DEBUG (console) — activable via env AL_DEBUG=1
# ======================================================
def debug_enabled() -> bool:
    try:
        return str(os.environ.get("AL_DEBUG", "1")).strip() not in ("0", "false", "False", "")
    except Exception:
        return True

def debug_print(*args, **kwargs):
    if not debug_enabled():
        return
    try:
        import datetime as _dt
        ts = _dt.datetime.now().strftime("%H:%M:%S")
        print("🐞", ts, *args, **kwargs, flush=True)
    except Exception:
        pass
import os
import requests
import json
from io import BytesIO
from openpyxl import load_workbook

def get_dropbox_access_token():
    r = requests.post(
        "https://api.dropbox.com/oauth2/token",
        data={
            "grant_type": "refresh_token",
            "refresh_token": os.environ["DROPBOX_REFRESH_TOKEN"],
            "client_id": os.environ["DROPBOX_APP_KEY"],
            "client_secret": os.environ["DROPBOX_APP_SECRET"],
        },
        timeout=10,
    )
    r.raise_for_status()
    return r.json()["access_token"]


def download_dropbox_excel_bytes(path="/Goldenlines/Planning 2026.xlsm"):
    token = get_dropbox_access_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Dropbox-API-Arg": f'{{"path": "{path}"}}',
        "Content-Type": "application/octet-stream",
    }
    r = requests.post(
        "https://content.dropboxapi.com/2/files/download",
        headers=headers,
        timeout=30,
    )
    r.raise_for_status()
    return r.content


def upload_dropbox_excel_bytes(content: bytes, path="/Goldenlines/Planning 2026.xlsm"):
    """Upload (overwrite) du fichier Excel vers Dropbox."""
    token = get_dropbox_access_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/octet-stream",
        "Dropbox-API-Arg": json.dumps({
            "path": path,
            "mode": "overwrite",
            "autorename": False,
            "mute": True,
            "strict_conflict": False,
        }),
    }
    r = requests.post(
        "https://content.dropboxapi.com/2/files/upload",
        headers=headers,
        data=content,
        timeout=60,
    )
    r.raise_for_status()
    return r.json()

@st.cache_data
def get_dropbox_excel_cached():
    return download_dropbox_excel_bytes()

def _cell_is_yellow(cell) -> bool:
    """
    Détecte le jaune Excel (fill, theme, indexed).
    Compatible Excel réel.
    """
    try:
        fill = cell.fill
        if fill is None or fill.patternType is None:
            return False

        fg = fill.fgColor
        if fg is None:
            return False

        # RGB direct
        if fg.type == "rgb" and fg.rgb:
            rgb = fg.rgb.upper()
            return rgb.endswith("FFFF00") or rgb in {"FFFFFF00", "00FFFF00"}

        # Indexed color (Excel ancien)
        if fg.type == "indexed":
            return fg.indexed in {5, 6}

        # Theme color (Excel moderne)
        if fg.type == "theme":
            return True

        return False
    except Exception:
        return False
GREEN_RGBS  = {"FF00B050", "FF92D050"}
ORANGE_RGBS = {"FFFFC000", "FFF4B084"}

def _cell_is_green(cell) -> bool:
    try:
        if not cell or not cell.fill or not cell.fill.fgColor:
            return False
        fg = cell.fill.fgColor
        if fg.type == "rgb" and fg.rgb:
            return fg.rgb.upper() in GREEN_RGBS
        if fg.type == "indexed":
            return fg.indexed in {17}
        if fg.type == "theme":
            return True
        return False
    except Exception:
        return False

def _cell_is_orange(cell) -> bool:
    try:
        if not cell or not cell.fill or not cell.fill.fgColor:
            return False
        fg = cell.fill.fgColor
        if fg.type == "rgb" and fg.rgb:
            return fg.rgb.upper() in ORANGE_RGBS
        if fg.type == "indexed":
            return fg.indexed in {45}
        return False
    except Exception:
        return False
import re

def parse_mail_to_navette(text: str) -> dict:
    """
    Transforme un mail brut en données navette (heuristique simple).
    """
    if not text:
        return {}

    t = text.lower()
    data = {}

    # 📆 Date
    m = re.search(r"(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})", text)
    if m:
        data["DATE"] = m.group(1)

    # ⏱ Heure
    m = re.search(r"(\d{1,2}[:h]\d{2})", text)
    if m:
        data["HEURE"] = m.group(1).replace("h", ":")

    # 👥 Pax
    m = re.search(r"(\d+)\s*(pax|personne|personnes)", t)
    if m:
        data["PAX"] = int(m.group(1))

    # ✈️ Vol
    m = re.search(r"\b([A-Z]{2}\s?\d{2,4})\b", text)
    if m:
        data["VOL"] = m.group(1).replace(" ", "")

    # 📍 Adresse simple
    m = re.search(r"\b\d{4}\s+[a-zà-ÿ\- ]+", t)
    if m:
        data["ADRESSE"] = m.group(0).title()

    # 🎯 Destination (règles simples)
    if "zaventem" in t or "bruxelles" in t:
        data["DESTINATION"] = "BRU"
    elif "charleroi" in t:
        data["DESTINATION"] = "CRL"
    elif "luxembourg" in t:
        data["DESTINATION"] = "LUX"

    data["RAW"] = text
    return data

# ======================================================
# 📘 FLAGS COULEURS EXCEL (DROPBOX)
# ======================================================

def add_excel_color_flags_from_dropbox(
    df: pd.DataFrame,
    sheet_name: str = "Feuil1"
) -> pd.DataFrame:

    df = df.copy().reset_index(drop=True)

    try:
        content = get_dropbox_excel_cached()
        if not content:
            raise RuntimeError("Fichier Dropbox inaccessible")

        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb[sheet_name]

        # Header Excel en ligne 2
        headers = [str(c.value).strip() if c.value else "" for c in ws[2]]

        def col_idx(name: str):
            name = name.strip().upper()
            for i, h in enumerate(headers):
                if h.upper() == name:
                    return i + 1
            return None

        col_date   = col_idx("DATE")
        col_heure  = col_idx("HEURE")
        col_ch     = col_idx("CH") or col_idx("CHAUFFEUR")
        col_caisse = col_idx("CAISSE") or col_idx("Caisse") or col_idx("PAIEMENT")

        is_groupage = []
        is_partage  = []
        is_paye     = []
        ack_excel   = []
        is_modif    = []

        # ======================================================
        # 🎨 LECTURE LIGNE PAR LIGNE
        # ======================================================
        for excel_row in range(3, 3 + len(df)):

            c_date   = ws.cell(excel_row, col_date)   if col_date else None
            c_heure  = ws.cell(excel_row, col_heure)  if col_heure else None
            c_ch     = ws.cell(excel_row, col_ch)     if col_ch else None
            c_caisse = ws.cell(excel_row, col_caisse) if col_caisse else None

            # 🟡 GROUPAGE / PARTAGE
            date_y  = _cell_is_yellow(c_date)  if c_date else False
            heure_y = _cell_is_yellow(c_heure) if c_heure else False

            is_groupage.append(1 if date_y and heure_y else 0)
            is_partage.append(1 if (not date_y) and heure_y else 0)

            # 💰 PAIEMENT
            is_paye.append(1 if c_caisse and _cell_is_green(c_caisse) else 0)

            # 👨‍✈️ CHAUFFEUR (Excel)
            if c_ch and _cell_is_green(c_ch):
                ack_excel.append(1)
                is_modif.append(0)
            elif c_ch and _cell_is_orange(c_ch):
                ack_excel.append(0)
                is_modif.append(1)
            else:
                ack_excel.append(0)
                is_modif.append(0)

        df["IS_GROUPAGE"] = is_groupage
        df["IS_PARTAGE"]  = is_partage
        df["IS_PAYE"]     = is_paye
        df["ACK_EXCEL"]   = ack_excel
        df["IS_MODIF"]    = is_modif

        # ⭐ ATTENTE (étoile chauffeur)
        if "CH" in df.columns:
            df["IS_ATTENTE"] = (
                df["CH"]
                .astype(str)
                .str.contains(r"\*", na=False)
                .astype(int)
            )
        else:
            df["IS_ATTENTE"] = 0

        return df

    except Exception as e:
        for col in ["IS_GROUPAGE", "IS_PARTAGE", "IS_ATTENTE", "IS_PAYE", "ACK_EXCEL", "IS_MODIF"]:
            df[col] = 0
        st.error(f"❌ Couleurs Excel non lues : {e}")
        return df


# ======================================================
# 🧾 LOGS (mémoire session) — visible dans l'UI
# ======================================================

import datetime as _dt


def log_event(message: str, level: str = "INFO"):
    """Ajoute une ligne de log en mémoire (st.session_state)."""
    try:
        ts = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] [{level.upper()}] {message}"
        try:
            debug_print(line)
        except Exception:
            pass
        if "logs" not in st.session_state:
            st.session_state["logs"] = []
        st.session_state["logs"].append(line)
        # limiter taille
        if len(st.session_state["logs"]) > 800:
            st.session_state["logs"] = st.session_state["logs"][-800:]
    except Exception:
        pass


def clear_logs():
    try:
        st.session_state["logs"] = []
    except Exception:
        pass


def render_logs_ui(title: str = "🧾 Logs", height: int = 260):
    """Affiche les logs dans Streamlit (safe)."""
    try:
        st.markdown(f"#### {title}")
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("🧹 Vider les logs", key="btn_clear_logs"):
                clear_logs()
        with col2:
            if st.button("🔄 Rafraîchir", key="btn_refresh_logs"):
                st.rerun()
        logs = st.session_state.get("logs") or []
        st.code("\n".join(logs), language="text")
    except Exception:
        pass



def _make_row_key_like_db(row: dict) -> str:
    """Même algorithme que database.make_row_key_from_row (copié ici pour éviter import circulaire)."""
    parts = [
        str(row.get("DATE", "") or ""),
        str(row.get("HEURE", "") or ""),
        str(row.get("Num BDC", "") or row.get("NUM BDC", "") or ""),
        str(row.get("NOM", "") or ""),
        str(row.get("ADRESSE", "") or ""),
        str(row.get("CP", "") or ""),
        str(row.get("Localité", "") or row.get("LOCALITE", "") or ""),
        str(row.get("Unnamed: 8", "") or ""),   # route
        str(row.get("DESIGNATION", "") or ""),
        str(row.get("VOL", "") or row.get("Vol", "") or ""),
    ]
    raw = "|".join(p.strip().upper() for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


# ============================================================
#   🆔 ROW_KEY (UUID) — Excel (Dropbox)
# ============================================================

def ensure_excel_row_key_column(
    dropbox_path: str = "/Goldenlines/Planning 2026.xlsm",
    sheet_name: str = "Feuil1",
    header_row: int = 2,
    data_start_row: int = 3,
    target_col_letter: str = "ZX",
) -> bool:
    """Assure la présence d'une colonne ROW_KEY (UUID) dans Excel.
    - La colonne est placée en ZX (par défaut) pour éviter toute mauvaise manip.
    - La colonne est masquée.
    - Les cellules vides reçoivent un UUID.
    Retourne True si le fichier a été modifié (upload), sinon False.
    """
    try:
        content = download_dropbox_excel_bytes(dropbox_path)
        if not content:
            return False
        wb = load_workbook(BytesIO(content))
        ws = wb[sheet_name]

        target_col = column_index_from_string(target_col_letter)

        # S'assurer que la feuille a au moins target_col colonnes (sinon openpyxl étend automatiquement)
        # On écrit l'en-tête + valeurs directement à l'index cible.
        header_cell = ws.cell(header_row, target_col)
        if str(header_cell.value or "").strip().upper() != "ROW_KEY":
            header_cell.value = "ROW_KEY"

        # Masquer la colonne
        ws.column_dimensions[target_col_letter].hidden = True

        changed = False
        import uuid

        for r in range(data_start_row, ws.max_row + 1):
            c = ws.cell(r, target_col)
            if c.value is None or str(c.value).strip() == "":
                c.value = str(uuid.uuid4())
                changed = True

        if changed:
            bio = BytesIO()
            wb.save(bio)
            upload_dropbox_excel_bytes(bio.getvalue(), dropbox_path)
        return changed
    except Exception as e:
        print(f"⚠️ ensure_excel_row_key_column error: {e}", flush=True)
        return False


def update_excel_rows_by_row_key(
    updates_by_row_key: dict,
    sheet_name: str = "Feuil1",
    dropbox_path: str = "/Goldenlines/Planning 2026.xlsm",
    row_key_col_letter: str = "ZX",
    debug: bool = False,
) -> int:
    """Met à jour des cellules dans Excel (Dropbox) en retrouvant les lignes via la colonne ROW_KEY.

    - La colonne ROW_KEY est attendue en ZX (par défaut) et masquée.
    - Ne recalcule plus le row_key depuis DATE/HEURE (trop fragile).
    """
    if not updates_by_row_key:
        return 0

    # 🔒 Assure la présence de la colonne ROW_KEY (UUID)
    ensure_excel_row_key_column(
        dropbox_path=dropbox_path,
        sheet_name=sheet_name,
        target_col_letter=row_key_col_letter,
    )

    content = download_dropbox_excel_bytes(dropbox_path)
    if not content:
        return 0

    wb = load_workbook(BytesIO(content))
    ws = wb[sheet_name]

    headers = [str(c.value).strip() if c.value else "" for c in ws[2]]
    col_map = {h.upper(): i + 1 for i, h in enumerate(headers) if h}

    # Map colonne ROW_KEY via lettre (ZX) : plus fiable que le header si Excel bouge
    from openpyxl.utils import column_index_from_string
    row_key_col = column_index_from_string(row_key_col_letter)

    # Index Excel : row_key -> excel_row
    excel_index = {}
    for excel_row in range(3, ws.max_row + 1):
        rk = ws.cell(excel_row, row_key_col).value
        if rk is None:
            continue
        rk = str(rk).strip()
        if rk:
            excel_index[rk] = excel_row

    # Debug demandé
    if debug:
        try:
            print("ROW_KEYS DB:", list(updates_by_row_key.keys())[:5])
            print("ROW_KEYS EXCEL:", list(excel_index.keys())[:5])
        except Exception:
            pass

    updated = 0
    for rk, updates in updates_by_row_key.items():
        excel_row = excel_index.get(str(rk).strip())
        if not excel_row:
            continue
        for k, v in (updates or {}).items():
            col_i = col_map.get(str(k).strip().upper())
            if not col_i:
                continue
            ws.cell(excel_row, col_i).value = v
        updated += 1

    if updated:
        bio = BytesIO()
        wb.save(bio)
        upload_dropbox_excel_bytes(bio.getvalue(), dropbox_path)

    return updated

def mark_caisse_paid_in_excel(row_keys: list, sheet_name: str = "Feuil1", dropbox_path: str = "/Goldenlines/Planning 2026.xlsm") -> int:
    """Met la cellule CAISSE en vert (payé) pour les lignes (row_key)."""
    if not row_keys:
        return 0
    from openpyxl.styles import PatternFill

    green_fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")

    content = download_dropbox_excel_bytes(dropbox_path)
    wb = load_workbook(BytesIO(content))
    ws = wb[sheet_name]

    headers = [str(c.value).strip() if c.value else "" for c in ws[2]]
    col_map = {h.upper(): i+1 for i, h in enumerate(headers) if h}
    caisse_col = col_map.get("CAISSE") or col_map.get("Caisse".upper())
    if not caisse_col:
        raise RuntimeError("Colonne CAISSE introuvable dans Excel")

    targets = set(str(k) for k in row_keys)

    updated = 0
    for excel_row in range(3, ws.max_row + 1):
        row_dict = {h: ws.cell(excel_row, col_i).value for h, col_i in col_map.items()}
        row_key = _make_row_key_like_db(row_dict)
        if row_key in targets:
            ws.cell(excel_row, caisse_col).fill = green_fill
            updated += 1

    if updated:
        bio = BytesIO()
        wb.save(bio)
        upload_dropbox_excel_bytes(bio.getvalue(), dropbox_path)

    return updated
# ============================================================
# 📥 MAIL → NAVETTE (V2) — parsing structuré
# ============================================================

def parse_mail_to_navette_v2(text: str) -> dict:
    """Parse un contenu copié-collé vers une structure exploitable par l'UI.

    Supporte 2 formats :
    1) Mail "structuré" (SECTEUR, Demandeur, Date/Heure/Pickup/Dest, ...)
    2) Tableau Excel copié-collé (TSV : colonnes séparées par des TAB)

    Retour :
      - mode "MAIL"  -> {meta..., "TRANSFERS": [ {DATE, HEURE, PICKUP, DEST, ...}, ... ]}
      - mode "TABLE" -> {"MODE":"TABLE", "ROWS": [ {<col>:<val>, ...}, ... ]}
    """
    if not text:
        return {}

    raw = str(text).strip("\ufeff\n\r\t ")
    if not raw:
        return {}

    # ------------------------------------------------------------
    # 1) Détection TABLE (copie Excel) : présence de TAB + header DATE/HEURE
    # ------------------------------------------------------------
    try:
        first_line = raw.splitlines()[0]
    except Exception:
        first_line = ""

    if "\t" in first_line and ("DATE" in first_line.upper()) and ("HEURE" in first_line.upper()):
        try:
            import pandas as pd
            from io import StringIO

            df = pd.read_csv(
                StringIO(raw),
                sep="\t",
                dtype=str,
                keep_default_na=False,
                engine="python",
            )

            # Nettoyage colonnes vides / unnamed
            df = df.rename(columns={c: str(c).strip() for c in df.columns})
            drop_cols = [c for c in df.columns if str(c).strip().lower().startswith("unnamed")]
            if drop_cols:
                df = df.drop(columns=drop_cols)

            # Nettoyage lignes vides
            df = df.fillna("")
            if df.shape[0] == 0:
                return {}

            rows = df.to_dict(orient="records")
            # Normaliser clés (strip)
            norm_rows = []
            for r in rows:
                nr = {}
                for k, v in (r or {}).items():
                    kk = str(k).strip()
                    nr[kk] = "" if v is None else str(v).strip()
                # ignorer lignes sans DATE/HEURE/NOM
                if not (nr.get("DATE") or nr.get("HEURE") or nr.get("NOM")):
                    continue
                norm_rows.append(nr)

            return {
                "MODE": "TABLE",
                "ROWS": norm_rows,
            }
        except Exception:
            # On retombe sur le mode mail si lecture TSV rate
            pass

    # ------------------------------------------------------------
    # 2) Mode MAIL structuré (ancien)
    # ------------------------------------------------------------
    lines = [l.strip() for l in raw.splitlines() if l.strip()]

    def find_after(label):
        lab = str(label).lower()
        for i, l in enumerate(lines):
            if lab in l.lower():
                if ":" in l:
                    v = l.split(":", 1)[1].strip()
                    if v:
                        return v
                if i + 1 < len(lines):
                    return lines[i + 1].strip()
        return ""

    data = {
        "MODE": "MAIL",
        "SECTEUR": find_after("SECTEUR"),
        "SBU": find_after("S.B.U.") or find_after("SBU"),
        "DEMANDEUR": find_after("Demandeur"),
        "VOYAGEURS": find_after("Voyageur"),
        "TEL": find_after("GSM"),
        "IMPUTATION": find_after("Imputation"),
        "SOCIETE": find_after("Société à facturer"),
        "TVA": find_after("TVA"),
        "BDC": find_after("Communication"),
        "TRANSFERS": [],
    }

    # Détection des trajets (Date / Heure / Pick-up / Destination)
    current = {}
    for l in lines:
        ll = l.lower()
        if ll.startswith("date"):
            if current:
                data["TRANSFERS"].append(current)
                current = {}
            current["DATE"] = l.split(":", 1)[1].strip() if ":" in l else ""
        elif "heure" in ll:
            current["HEURE"] = l.split(":", 1)[1].strip() if ":" in l else ""
        elif "lieu de pick" in ll:
            current["PICKUP"] = l.split(":", 1)[1].strip() if ":" in l else ""
        elif "transfert" in ll and (" à" in ll or " a" in ll):
            current["DEST"] = l.split(":", 1)[1].strip() if ":" in l else ""

    if current:
        data["TRANSFERS"].append(current)

    return data

# ============================================================
# ⚡ CACHE parsing MAIL → NAVETTE (évite re-parse à chaque clic)
# ============================================================

@st.cache_data(ttl=3600, show_spinner=False)
def parse_mail_to_navette_v2_cached(text: str) -> dict:
    """Wrapper cache (1h) autour de parse_mail_to_navette_v2.
    Objectif : rendre l'UI fluide (checkbox urgences / edits) sans recalcul.
    """
    return parse_mail_to_navette_v2(text)

def format_mail_navette_v2(data: dict) -> str:
    """
    Génère le mail FINAL navette (copiable tel quel).
    """
    if not data:
        return ""

    lines = []
    lines.append("Bonjour,\n")
    lines.append(
        f"Ci-dessous ma demande de navette pour {data.get('VOYAGEURS','')}."
    )
    lines.append("Merci de l’ajouter à votre planning :\n")

    def block(label, value):
        if value:
            lines.append(f"{label} :\n{value}\n")

    block("SECTEUR", data.get("SECTEUR"))
    block("S.B.U.", data.get("SBU"))
    block("Demandeur", data.get("DEMANDEUR"))
    block("Voyageur(s)", data.get("VOYAGEURS"))
    block("N° GSM du voyageur", data.get("TEL"))
    block("Imputation", data.get("IMPUTATION"))
    block("Société à facturer", data.get("SOCIETE"))
    block("TVA", data.get("TVA"))
    block("Communication", data.get("BDC"))

    lines.append("\n")

    for t in data.get("TRANSFERS", []):
        lines.append("Date :")
        lines.append(t.get("DATE", ""))
        lines.append("\nHeure de pick up :")
        lines.append(t.get("HEURE", ""))
        lines.append("\nLieu de pick up :")
        lines.append(t.get("PICKUP", ""))
        lines.append("\nTransfert à :")
        lines.append(t.get("DEST", ""))
        lines.append("\n")

    lines.append(
        "Merci d’avance pour votre compréhension et la confirmation de cette course.\n"
    )
    lines.append("Cordialement,")

    return "\n".join(lines).strip()



# ============================================================
#   🧠 ALIAS & SUGGESTIONS — UTILITAIRES (sans import DB au chargement)
# ============================================================

def detect_dest_code(text: str) -> str:
    """Détecte un code destination (BRU/CRL/GUIL/JCO/...) à partir d'un texte libre."""
    s = (text or "").upper().strip()
    if not s:
        return ""

    # Synonymes fréquents (robuste aux mails)
    synonyms = {
        "ZAVENTEM": "BRU",
        "BRUSSELS AIRPORT": "BRU",
        "BRUXELLES AIRPORT": "BRU",
        "CHARLEROI": "CRL",
        "BRUSSELS SOUTH": "CRL",
        "GUILLEMINS": "GUIL",
        "GUIL": "GUIL",
        "PARIS NORD": "BXL_MIDI",
        "BXL MIDI": "BXL_MIDI",
        "BRUXELLES MIDI": "BXL_MIDI",
    }

    for k, v in synonyms.items():
        if k in s:
            return v

    # 1) Codes connus (table DB)
    try:
        from database import get_location_aliases_df
        df = get_location_aliases_df()
        if df is not None and not df.empty:
            # match code direct
            codes = (
                df["code"]
                .fillna("")
                .astype(str)
                .str.upper()
                .str.strip()
                .tolist()
            )
            for c in codes:
                if c and c in s:
                    return c

            # match label
            for _, r in df.iterrows():
                c = str(r.get("code", "") or "").upper().strip()
                lbl = str(r.get("label", "") or "").upper().strip()
                if c and lbl and lbl in s:
                    return c
    except Exception:
        pass

    return ""


def suggest_heures_from_rules(dest_text_or_code: str, sens: str = "VERS") -> str:
    """Retourne une suggestion d'heures (ex '2.5') selon time_rules."""
    dest = (dest_text_or_code or "").upper().strip()
    if not dest:
        return ""

    # si on reçoit un texte libre, tente d'abord de détecter un code
    code = detect_dest_code(dest) or dest

    try:
        from database import get_time_rules_df
        df = get_time_rules_df()
        if df is None or df.empty:
            return ""

        s = (sens or "").upper().strip()
        if not s:
            s = "VERS"

        # Filtre sens si colonne présente
        if "sens" in df.columns:
            df2 = df[df["sens"].fillna("").astype(str).str.upper().str.strip().isin(["*", s])].copy()
        else:
            df2 = df.copy()

        # match sur dest_contains
        if "dest_contains" in df2.columns:
            for _, r in df2.iterrows():
                dc = str(r.get("dest_contains", "") or "").upper().strip()
                if not dc or dc == "*":
                    continue
                if dc in code or dc in dest:
                    minutes = r.get("minutes", 0) or 0
                    try:
                        return str(round(float(minutes) / 60, 2))
                    except Exception:
                        return ""
        # fallback: règle wildcard
        if "minutes" in df2.columns:
            row = df2.head(1)
            if not row.empty:
                minutes = row.iloc[0].get("minutes", 0) or 0
                try:
                    return str(round(float(minutes) / 60, 2))
                except Exception:
                    return ""
    except Exception:
        pass

    return ""
import re
from datetime import date, datetime
import pandas as pd

def parse_mail_to_navette_flexible(text: str) -> dict:
    """
    Parseur tolérant :
    - accepte mails humains / listes
    - multi-dates
    - DE/VERS implicite
    """
    rows = []
    current_date = None

    lines = [l.strip() for l in text.splitlines() if l.strip()]

    for line in lines:

        # -------------------------
        # DATE (25/02, 25-02-2026…)
        # -------------------------
        m_date = re.search(r'(\d{1,2}[\/\-]\d{1,2}(?:[\/\-]\d{2,4})?)', line)
        if m_date:
            d = m_date.group(1)
            try:
                dt = pd.to_datetime(d, dayfirst=True)
                current_date = dt.strftime("%d/%m/%Y")
                continue
            except Exception:
                pass

        # -------------------------
        # HEURE (19H10, 07:30…)
        # -------------------------
        m_time = re.search(r'(\d{1,2})[:hH](\d{2})', line)
        if m_time and current_date:
            hh, mm = m_time.groups()
            heure = f"{int(hh):02d}:{int(mm):02d}"

            # destination implicite
            dest = "ZAV" if "ZAV" in line.upper() else ""

            rows.append({
                "DATE": current_date,
                "HEURE": heure,
                "CH": "",
                "NOM": "À confirmer",
                "DESIGNATION": dest,
                "ADRESSE": "",
                "Tél": "",
                "REMARQUE": line,
            })
            continue

        # -------------------------
        # GSM
        # -------------------------
        m_gsm = re.search(r'(0\d{3}[\s\.]?\d{2}[\s\.]?\d{2}[\s\.]?\d{2})', line)
        if m_gsm and rows:
            rows[-1]["Tél"] = m_gsm.group(1)

        # -------------------------
        # Adresse (heuristique simple)
        # -------------------------
        if any(x in line.lower() for x in ["rue", "avenue", "chaussée", "place"]):
            if rows:
                rows[-1]["ADRESSE"] = line

    return {
        "SOURCE": "flexible",
        "ROWS": rows,
    }


# ======================================================
# 📘 FLAGS FACTURE + CLIENT HUB
# ======================================================

def read_client_hub_excel_flags(sheet_name: str = "Feuil1") -> pd.DataFrame:
    """Retourne un DataFrame léger avec row-index Excel et flags client:
    - FACTURE_ENVOYEE si P+Q verts
    - BDC_NOM_VERT si P+Q verts
    """
    out = []
    try:
        content = get_dropbox_excel_cached()
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb[sheet_name]
        # colonnes fixes Feuil1: P=16, Q=17
        for excel_row in range(3, ws.max_row + 1):
            c_p = ws.cell(excel_row, 16)
            c_q = ws.cell(excel_row, 17)
            fact = 1 if _cell_is_green(c_p) and _cell_is_green(c_q) else 0
            out.append({
                "_excel_row": excel_row,
                "FACTURE_ENVOYEE": fact,
                "BDC_NOM_VERT": fact,
            })
    except Exception:
        pass
    return pd.DataFrame(out)

def read_dates_carburant_map() -> dict:
    try:
        content = get_dropbox_excel_cached()
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb["dates carburant"]
        mapping = {}
        for r in range(2, ws.max_row + 1):
            d = ws.cell(r, 1).value
            coef = ws.cell(r, 3).value
            if not d:
                continue
            try:
                key = pd.to_datetime(d, errors="coerce").strftime("%Y-%m-%d")
            except Exception:
                continue
            try:
                mapping[key] = float(coef or 0)
            except Exception:
                mapping[key] = 0.0
        return mapping
    except Exception:
        return {}

def compute_surcharge_like_xlsm(date_iso: str, km, htva, coef_map: dict) -> float:
    try:
        if not date_iso:
            return 0.0
        d = pd.to_datetime(date_iso, errors="coerce")
        if pd.isna(d):
            return 0.0
        if d.date() < pd.Timestamp("2026-04-01").date():
            return 0.0
        coef = float(coef_map.get(d.strftime("%Y-%m-%d"), 0) or 0)
        if coef <= 0:
            return 0.0
        h = float(htva or 0)
        if h in (47.5, 55.0):
            return 2.0
        if h in (115.0, 148.5):
            return round(200.0 * coef, 2)
        k = float(km or 0)
        if k > 0:
            return round(k * coef, 2)
        return 0.0
    except Exception:
        return 0.0
