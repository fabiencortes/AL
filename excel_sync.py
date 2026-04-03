# excel_sync.py
# ============================================================
# 🔁 SYNC DB → EXCEL (MANUEL, STABLE)
# - Headers en ligne 2 (auto-détection)
# - Matching par row_key SHA1 (database.make_row_key_from_row)
# ============================================================

from __future__ import annotations

from typing import Dict, Any, Tuple, Optional
import openpyxl

from database import make_row_key_from_row


def _detect_header_row(ws, max_scan_rows: int = 10) -> int:
    """
    Cherche la ligne d'en-tête dans les premières lignes.
    On considère que c'est une ligne header si elle contient au moins DATE et HEURE.
    """
    need = {"DATE", "HEURE"}
    best_row = 2  # par défaut (ton fichier Planning 2026.xlsx)
    for r in range(1, max_scan_rows + 1):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip().upper()
            if s:
                vals.append(s)
        if not vals:
            continue
        if need.issubset(set(vals)):
            best_row = r
            break
    return best_row


def _read_headers(ws, header_row: int) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        name = str(v).strip()
        if not name:
            continue
        headers[name.strip().upper()] = c
    return headers


def _row_to_dict(ws, r: int, headers: Dict[str, int], header_row: int) -> Dict[str, Any]:
    row: Dict[str, Any] = {}
    for name_u, c in headers.items():
        row[name_u] = ws.cell(row=r, column=c).value
    # fournir un index de ligne pour éviter collisions si ligne "pauvre"
    row["__excel_row"] = r
    return row


def build_rowkey_to_rownum_map(xlsx_path: str, sheet_name: str = "Feuil1") -> Tuple[Dict[str, int], Dict[str, int], int]:
    """
    Retourne:
      - map row_key_sha1 -> row_number Excel
      - headers map (header_name_upper -> column index)
      - header_row index
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    header_row = _detect_header_row(ws)
    headers = _read_headers(ws, header_row)

    # Colonnes minimales
    if "DATE" not in headers or "HEURE" not in headers:
        raise ValueError(f"Headers Excel introuvables (DATE/HEURE). Ligne header détectée: {header_row}")

    mapping: Dict[str, int] = {}

    for r in range(header_row + 1, ws.max_row + 1):
        # ignorer lignes vides
        if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, ws.max_column + 1)):
            continue
        row_dict = _row_to_dict(ws, r, headers, header_row)

        # database.make_row_key_from_row attend des clés originales (DATE, HEURE, etc.)
        # Ici on a les headers upper => adapter quelques noms
        normalized = {}
        for k_u, v in row_dict.items():
            # remettre quelques clés attendues
            if k_u == "NUM BDC":
                normalized["Num BDC"] = v
            elif k_u == "LOCALITÉ" or k_u == "LOCALITE":
                normalized["Localité"] = v
            else:
                # garder en version "DATE", "HEURE", "NOM", etc.
                normalized[k_u] = v
        # Ajoute le fallback
        normalized["__excel_row"] = row_dict.get("__excel_row")

        rk = make_row_key_from_row(normalized)
        # en cas de duplicat (rare), on garde la première occurrence
        mapping.setdefault(rk, r)

    return mapping, headers, header_row


def apply_updates_to_excel(
    xlsx_path: str,
    updates: Dict[str, Dict[str, Any]],
    sheet_name: str = "Feuil1",
) -> int:
    """
    updates = {row_key_sha1: {"CH": "FA", ...}}
    Applique uniquement les colonnes supportées si elles existent dans l'Excel.
    """
    if not updates:
        return 0

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    header_row = _detect_header_row(ws)
    headers = _read_headers(ws, header_row)

    rk_map, _, _ = build_rowkey_to_rownum_map(xlsx_path, sheet_name=sheet_name)

    applied = 0

    for rk, patch in updates.items():
        r = rk_map.get(rk)
        if not r:
            continue
        for col_name, val in patch.items():
            col_u = str(col_name).strip().upper()
            if col_u not in headers:
                # colonne absente => on ignore
                continue
            c = headers[col_u]
            ws.cell(row=r, column=c).value = val
        applied += 1

    wb.save(xlsx_path)
    return applied