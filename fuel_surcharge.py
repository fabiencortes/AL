
from __future__ import annotations

import sqlite3
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

DB_NAME = "airportslines.db"
DROPBOX_PATH = "/Goldenlines/Planning 2026.xlsx"
DEFAULT_BASE_DIESEL = 1.54
DEFAULT_CURRENT_DIESEL = 2.30
DEFAULT_CONSO_100 = 8.0
DEFAULT_RECOVERY = 100.0
DEFAULT_XLSX_HEADER = "Surcharge carburant"
DEFAULT_XLSX_COL_LETTER = "AA"
DEFAULT_CLIENTS = [
    {"Client": "Tarif 0.55", "Prix_km": 0.55, "Recuperation_pct": 100.0, "Actif": True, "Match": "", "Notes": ""},
    {"Client": "Tarif 0.60", "Prix_km": 0.60, "Recuperation_pct": 100.0, "Actif": True, "Match": "", "Notes": ""},
    {"Client": "Tarif 0.70", "Prix_km": 0.70, "Recuperation_pct": 100.0, "Actif": True, "Match": "", "Notes": ""},
    {"Client": "Tarif 0.78", "Prix_km": 0.78, "Recuperation_pct": 100.0, "Actif": True, "Match": "", "Notes": ""},
]


# ---------------------------------------------------------
# DB helpers
# ---------------------------------------------------------
def _get_db_path() -> Path:
    return Path.cwd() / DB_NAME


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(_get_db_path(), timeout=60)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout=5000")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn


def _ensure_column(conn: sqlite3.Connection, table: str, column: str, definition: str) -> None:
    cols = {r['name'] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in cols:
        conn.execute(f'ALTER TABLE {table} ADD COLUMN "{column}" {definition}')


def _init_tables() -> None:
    with _connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS fuel_surcharge_clients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client_name TEXT NOT NULL UNIQUE,
                match_text TEXT DEFAULT '',
                price_per_km REAL NOT NULL DEFAULT 0,
                recovery_pct REAL NOT NULL DEFAULT 100,
                active INTEGER NOT NULL DEFAULT 1,
                notes TEXT DEFAULT '',
                updated_at TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS fuel_surcharge_settings (
                setting_key TEXT PRIMARY KEY,
                setting_value TEXT
            )
            """
        )

        # Migration douce si une ancienne table existe déjà sans certaines colonnes
        _ensure_column(conn, "fuel_surcharge_clients", "match_text", "TEXT DEFAULT ''")
        _ensure_column(conn, "fuel_surcharge_clients", "price_per_km", "REAL NOT NULL DEFAULT 0")
        _ensure_column(conn, "fuel_surcharge_clients", "recovery_pct", "REAL NOT NULL DEFAULT 100")
        _ensure_column(conn, "fuel_surcharge_clients", "active", "INTEGER NOT NULL DEFAULT 1")
        _ensure_column(conn, "fuel_surcharge_clients", "notes", "TEXT DEFAULT ''")
        _ensure_column(conn, "fuel_surcharge_clients", "updated_at", "TEXT")

        cols = {r['name'] for r in conn.execute("PRAGMA table_info(planning)")}
        if "SURCHARGE_CARBURANT" not in cols:
            conn.execute('ALTER TABLE planning ADD COLUMN "SURCHARGE_CARBURANT" REAL DEFAULT 0')
        if "SURCHARGE_CARBURANT_AT" not in cols:
            conn.execute('ALTER TABLE planning ADD COLUMN "SURCHARGE_CARBURANT_AT" TEXT')
        if "SURCHARGE_SOURCE" not in cols:
            conn.execute('ALTER TABLE planning ADD COLUMN "SURCHARGE_SOURCE" TEXT')
        conn.commit()

    if _load_clients_from_db().empty:
        _save_clients_to_db(pd.DataFrame(DEFAULT_CLIENTS))

    defaults = {
        "base_diesel": str(DEFAULT_BASE_DIESEL),
        "current_diesel": str(DEFAULT_CURRENT_DIESEL),
        "conso_100": str(DEFAULT_CONSO_100),
        "recovery_default": str(DEFAULT_RECOVERY),
        "xlsx_header": DEFAULT_XLSX_HEADER,
        "xlsx_col_letter": DEFAULT_XLSX_COL_LETTER,
    }
    settings = _load_settings_from_db()
    changed = False
    for k, v in defaults.items():
        if k not in settings:
            settings[k] = v
            changed = True
    if changed:
        _save_settings_to_db(settings)


def _load_settings_from_db() -> dict[str, str]:
    with _connect() as conn:
        rows = conn.execute("SELECT setting_key, setting_value FROM fuel_surcharge_settings").fetchall()
    return {str(r['setting_key']): str(r['setting_value'] or '') for r in rows}


def _save_settings_to_db(settings: dict[str, str]) -> None:
    with _connect() as conn:
        for k, v in settings.items():
            conn.execute(
                """
                INSERT INTO fuel_surcharge_settings(setting_key, setting_value)
                VALUES(?, ?)
                ON CONFLICT(setting_key) DO UPDATE SET setting_value = excluded.setting_value
                """,
                (str(k), str(v)),
            )
        conn.commit()


def _load_clients_from_db() -> pd.DataFrame:
    with _connect() as conn:
        cols = {r['name'] for r in conn.execute("PRAGMA table_info(fuel_surcharge_clients)").fetchall()}
        select_parts = [
            "client_name AS Client",
            "price_per_km AS Prix_km",
            "recovery_pct AS Recuperation_pct",
            "CASE WHEN active = 1 THEN 1 ELSE 0 END AS Actif",
        ]
        if "match_text" in cols:
            select_parts.insert(1, "match_text AS Match")
        else:
            select_parts.insert(1, "'' AS Match")
        if "notes" in cols:
            select_parts.append("COALESCE(notes, '') AS Notes")
        else:
            select_parts.append("'' AS Notes")
        sql = f"""
            SELECT
                {', '.join(select_parts)}
            FROM fuel_surcharge_clients
            ORDER BY client_name COLLATE NOCASE
        """
        rows = conn.execute(sql).fetchall()
    if not rows:
        return pd.DataFrame(columns=["Client", "Match", "Prix_km", "Recuperation_pct", "Actif", "Notes"])
    df = pd.DataFrame([dict(r) for r in rows])
    df["Actif"] = df["Actif"].astype(bool)
    return df


def _save_clients_to_db(df: pd.DataFrame) -> None:
    clean = df.copy() if df is not None else pd.DataFrame()
    if clean.empty:
        clean = pd.DataFrame(columns=["Client", "Match", "Prix_km", "Recuperation_pct", "Actif", "Notes"])

    clean = clean.fillna("")
    for c in ["Client", "Match", "Notes"]:
        if c in clean.columns:
            clean[c] = clean[c].astype(str).str.strip()
    clean = clean[clean["Client"] != ""]

    if not clean.empty:
        clean["Prix_km"] = pd.to_numeric(clean["Prix_km"], errors="coerce").fillna(0.0)
        clean["Recuperation_pct"] = pd.to_numeric(clean["Recuperation_pct"], errors="coerce").fillna(DEFAULT_RECOVERY).clip(lower=0, upper=200)
        clean["Actif"] = clean["Actif"].astype(bool)
        clean = clean.drop_duplicates(subset=["Client"], keep="first")

    now = datetime.now().isoformat(timespec="seconds")
    with _connect() as conn:
        conn.execute("DELETE FROM fuel_surcharge_clients")
        for _, row in clean.iterrows():
            conn.execute(
                """
                INSERT INTO fuel_surcharge_clients(client_name, match_text, price_per_km, recovery_pct, active, notes, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["Client"],
                    row.get("Match", ""),
                    float(row["Prix_km"]),
                    float(row["Recuperation_pct"]),
                    1 if bool(row["Actif"]) else 0,
                    row.get("Notes", ""),
                    now,
                ),
            )
        conn.commit()


# ---------------------------------------------------------
# Business helpers
# ---------------------------------------------------------
def _compute_metrics(km: float, price_per_km: float, diesel_base: float, diesel_current: float, conso_100: float, recovery_pct: float) -> dict[str, float]:
    km = max(float(km or 0), 0.0)
    price_per_km = max(float(price_per_km or 0), 0.0)
    diesel_base = max(float(diesel_base or 0), 0.0)
    diesel_current = max(float(diesel_current or 0), 0.0)
    conso_100 = max(float(conso_100 or 0), 0.0)
    recovery_pct = max(float(recovery_pct or 0), 0.0)

    conso_l_km = conso_100 / 100.0
    cost_km_base = diesel_base * conso_l_km
    cost_km_current = diesel_current * conso_l_km
    raw_surcharge_km = max(cost_km_current - cost_km_base, 0.0)
    applied_surcharge_km = raw_surcharge_km * (recovery_pct / 100.0)
    fuel_total_current = km * cost_km_current
    surcharge_total = km * applied_surcharge_km
    base_trip_total = km * price_per_km
    price_share_pct = (cost_km_current / price_per_km * 100.0) if price_per_km > 0 else 0.0
    surcharge_pct = (applied_surcharge_km / price_per_km * 100.0) if price_per_km > 0 else 0.0
    return {
        "cost_km_base": cost_km_base,
        "cost_km_current": cost_km_current,
        "raw_surcharge_km": raw_surcharge_km,
        "applied_surcharge_km": applied_surcharge_km,
        "fuel_total_current": fuel_total_current,
        "surcharge_total": surcharge_total,
        "price_share_pct": price_share_pct,
        "surcharge_pct": surcharge_pct,
        "recommended_price_km": price_per_km + applied_surcharge_km,
        "trip_total_without_surcharge": base_trip_total,
        "trip_total_with_surcharge": base_trip_total + surcharge_total,
    }


def _fmt_eur(v: float, digits: int = 4) -> str:
    return f"{float(v):,.{digits}f} €".replace(",", "X").replace(".", ",").replace("X", " ")


def _fmt_pct(v: float) -> str:
    return f"{float(v):,.2f} %".replace(",", "X").replace(".", ",").replace("X", " ")


def _to_float(val) -> float:
    try:
        return float(str(val).replace(",", ".").strip())
    except Exception:
        return 0.0


def _effective_km(row: pd.Series) -> float:
    for key in ["KM", "KM_EST"]:
        if key in row and str(row.get(key, "")).strip() not in ("", "nan", "None"):
            v = _to_float(row.get(key))
            if v > 0:
                return v
    return 0.0


def _load_planning_rows(start_date: date, end_date: date, only_active: bool = True) -> pd.DataFrame:
    sql = """
        SELECT *
        FROM planning
        WHERE date(COALESCE(DATE_ISO, '')) BETWEEN date(?) AND date(?)
    """
    if only_active:
        sql += " AND COALESCE(IS_INDISPO,0)=0 AND COALESCE(IS_SUPERSEDED,0)=0"
    sql += " ORDER BY DATE_ISO, HEURE, id"
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn, params=(start_date.isoformat(), end_date.isoformat()))
    if df.empty:
        return df
    for col in ["NOM", "DEMANDEUR", "Num BDC", "GO", "CH", "DESIGNATION", "SURCHARGE_CARBURANT"]:
        if col not in df.columns:
            df[col] = ""
    df["KM_EFFECTIF"] = df.apply(_effective_km, axis=1)
    return df


def _find_matching_rule(row: pd.Series, df_clients: pd.DataFrame) -> Optional[pd.Series]:
    if df_clients is None or df_clients.empty:
        return None
    haystack = " | ".join([
        str(row.get("NOM", "") or ""),
        str(row.get("DEMANDEUR", "") or ""),
        str(row.get("Num BDC", "") or ""),
        str(row.get("GO", "") or ""),
        str(row.get("DESIGNATION", "") or ""),
    ]).lower()
    for _, rule in df_clients[df_clients["Actif"] == True].iterrows():
        match = str(rule.get("Match", "") or "").strip().lower()
        client_name = str(rule.get("Client", "") or "").strip().lower()
        if match and match in haystack:
            return rule
        if client_name and client_name not in {"tarif 0.55", "tarif 0.60", "tarif 0.70", "tarif 0.78"} and client_name in haystack:
            return rule
    return None


def _update_db_surcharge(row_id: int, amount: float, source: str) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with _connect() as conn:
        conn.execute(
            """
            UPDATE planning
            SET SURCHARGE_CARBURANT = ?,
                SURCHARGE_CARBURANT_AT = ?,
                SURCHARGE_SOURCE = ?
            WHERE id = ?
            """,
            (round(float(amount), 2), now, str(source), int(row_id)),
        )
        conn.commit()


# ---------------------------------------------------------
# Excel Dropbox helpers
# ---------------------------------------------------------
def _download_dropbox_excel_bytes() -> Optional[bytes]:
    try:
        from utils import download_dropbox_excel_bytes
        return download_dropbox_excel_bytes(DROPBOX_PATH)
    except Exception:
        return None


def _upload_dropbox_excel_bytes(content: bytes) -> bool:
    try:
        from utils import upload_dropbox_excel_bytes
        upload_dropbox_excel_bytes(content, DROPBOX_PATH)
        return True
    except Exception:
        return False


def _find_header_row(ws) -> Optional[int]:
    for i in range(1, min(15, ws.max_row) + 1):
        vals = [str(ws.cell(row=i, column=c).value or "").strip().upper() for c in range(1, min(80, ws.max_column) + 1)]
        if "DATE" in vals and "HEURE" in vals:
            return i
    return None


def _get_header_map(ws, header_row: int) -> dict[str, int]:
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            out[name.upper()] = c
    return out


def _ensure_excel_target_col(ws, header_row: int, header_name: str, preferred_letter: str = DEFAULT_XLSX_COL_LETTER) -> int:
    header_map = _get_header_map(ws, header_row)
    existing = header_map.get(str(header_name).strip().upper())
    if existing:
        return existing
    preferred_idx = ws[preferred_letter + str(header_row)].column if preferred_letter else ws.max_column + 1
    if preferred_idx <= ws.max_column:
        cur_header = str(ws.cell(row=header_row, column=preferred_idx).value or "").strip()
        if cur_header not in ("", header_name):
            preferred_idx = ws.max_column + 1
    ws.cell(row=header_row, column=preferred_idx).value = header_name
    return preferred_idx


def _write_surcharge_to_excel(row_key: str, amount: float, header_name: str, preferred_letter: str = DEFAULT_XLSX_COL_LETTER) -> tuple[bool, str]:
    content = _download_dropbox_excel_bytes()
    if not content:
        return False, "Dropbox inaccessible"
    wb = load_workbook(BytesIO(content))
    if "Feuil1" not in wb.sheetnames:
        return False, "Feuil1 introuvable"
    ws = wb["Feuil1"]
    header_row = _find_header_row(ws)
    if not header_row:
        return False, "En-tête Feuil1 introuvable"
    headers = _get_header_map(ws, header_row)
    row_key_col = headers.get("ROW_KEY") or headers.get("ROW KEY")
    if not row_key_col:
        return False, "Colonne ROW_KEY introuvable dans Excel"
    target_col = _ensure_excel_target_col(ws, header_row, header_name, preferred_letter)

    found = False
    for r in range(header_row + 1, ws.max_row + 1):
        rk = str(ws.cell(row=r, column=row_key_col).value or "").strip()
        if rk and rk == str(row_key).strip():
            ws.cell(row=r, column=target_col).value = round(float(amount), 2)
            found = True
            break

    if not found:
        return False, "Transfert introuvable dans Excel via ROW_KEY"

    out = BytesIO()
    wb.save(out)
    ok = _upload_dropbox_excel_bytes(out.getvalue())
    return ok, "OK" if ok else "Échec upload Dropbox"


# ---------------------------------------------------------
# UI helpers
# ---------------------------------------------------------
def _ensure_state() -> None:
    if "fuel_settings" not in st.session_state:
        st.session_state["fuel_settings"] = _load_settings_from_db()
    if "fuel_clients_df" not in st.session_state:
        st.session_state["fuel_clients_df"] = _load_clients_from_db()


def _reload_state() -> None:
    st.session_state["fuel_settings"] = _load_settings_from_db()
    st.session_state["fuel_clients_df"] = _load_clients_from_db()


def _transfer_label(row: pd.Series) -> str:
    km = _effective_km(row)
    return f"#{int(row.get('id', 0))} | {row.get('DATE','')} {row.get('HEURE','')} | {row.get('CH','')} | {row.get('NOM','')} | {row.get('DESIGNATION','')} | {km:.0f} km"


def render_fuel_tab() -> None:
    _init_tables()
    _ensure_state()

    settings = st.session_state["fuel_settings"]
    df_clients = st.session_state["fuel_clients_df"].copy()

    diesel_base = float(settings.get("base_diesel", DEFAULT_BASE_DIESEL) or DEFAULT_BASE_DIESEL)
    diesel_current = float(settings.get("current_diesel", DEFAULT_CURRENT_DIESEL) or DEFAULT_CURRENT_DIESEL)
    conso_100 = float(settings.get("conso_100", DEFAULT_CONSO_100) or DEFAULT_CONSO_100)
    recovery_default = float(settings.get("recovery_default", DEFAULT_RECOVERY) or DEFAULT_RECOVERY)
    xlsx_header = str(settings.get("xlsx_header", DEFAULT_XLSX_HEADER) or DEFAULT_XLSX_HEADER)
    xlsx_col_letter = str(settings.get("xlsx_col_letter", DEFAULT_XLSX_COL_LETTER) or DEFAULT_XLSX_COL_LETTER)

    st.subheader("⛽ Surcharge carburant — version PRO")
    st.caption("Calcul par trajet, règles client, écriture en DB et directement dans le XLSX Dropbox sur la bonne ligne.")

    with st.expander("ℹ️ Base de calcul", expanded=False):
        st.markdown(
            """
- **Base carburant** : diesel de base vs diesel actuel.
- **Carburant réel/km** = diesel actuel × (8 L / 100 km) ou la conso réglée.
- **Surcoût/km** = (diesel actuel - diesel de base) × conso/100.
- **Surcharge appliquée** = surcoût × % de récupération.
- **Base km** : le calcul prend d'abord `KM`, sinon `KM_EST`.
            """
        )
        st.warning("Pour un calcul juste, il vaut mieux remplir les bons kilomètres dans la colonne KM. Sinon le module prendra KM_EST quand il existe, mais ce sera seulement une estimation.")

    with st.form("fuel_settings_form_pro", border=True):
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            new_base = st.number_input("Diesel de base (€ / L)", min_value=0.0, value=diesel_base, step=0.01)
        with c2:
            new_current = st.number_input("Diesel actuel (€ / L)", min_value=0.0, value=diesel_current, step=0.01)
        with c3:
            new_conso = st.number_input("Consommation (L / 100 km)", min_value=0.0, value=conso_100, step=0.1)
        with c4:
            new_recovery = st.number_input("Récupération défaut (%)", min_value=0.0, max_value=200.0, value=recovery_default, step=5.0)
        c5, c6 = st.columns(2)
        with c5:
            new_header = st.text_input("Nom colonne XLSX", value=xlsx_header)
        with c6:
            new_letter = st.text_input("Colonne Excel préférée", value=xlsx_col_letter)
        if st.form_submit_button("💾 Enregistrer les paramètres", use_container_width=True):
            _save_settings_to_db({
                "base_diesel": str(new_base),
                "current_diesel": str(new_current),
                "conso_100": str(new_conso),
                "recovery_default": str(new_recovery),
                "xlsx_header": str(new_header).strip() or DEFAULT_XLSX_HEADER,
                "xlsx_col_letter": str(new_letter).strip().upper() or DEFAULT_XLSX_COL_LETTER,
            })
            _reload_state()
            st.success("Paramètres enregistrés.")
            st.rerun()

    diesel_base = float(st.session_state["fuel_settings"].get("base_diesel", DEFAULT_BASE_DIESEL))
    diesel_current = float(st.session_state["fuel_settings"].get("current_diesel", DEFAULT_CURRENT_DIESEL))
    conso_100 = float(st.session_state["fuel_settings"].get("conso_100", DEFAULT_CONSO_100))
    recovery_default = float(st.session_state["fuel_settings"].get("recovery_default", DEFAULT_RECOVERY))
    xlsx_header = str(st.session_state["fuel_settings"].get("xlsx_header", DEFAULT_XLSX_HEADER))
    xlsx_col_letter = str(st.session_state["fuel_settings"].get("xlsx_col_letter", DEFAULT_XLSX_COL_LETTER))

    base_metrics = _compute_metrics(1, 0.70, diesel_base, diesel_current, conso_100, 100.0)
    m1, m2, m3 = st.columns(3)
    m1.metric("Carburant actuel / km", _fmt_eur(base_metrics["cost_km_current"]))
    m2.metric("Surcoût réel / km", _fmt_eur(base_metrics["raw_surcharge_km"]))
    m3.metric("Carburant base / km", _fmt_eur(base_metrics["cost_km_base"]))

    tab_normal, tab_clients, tab_transfers, tab_auto = st.tabs(["Mode normal", "Clients / règles", "Transfert précis", "Auto jour / période"])

    with tab_normal:
        with st.form("fuel_normal_form_pro", border=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                km = st.number_input("Km du trajet", min_value=0.0, value=180.0, step=1.0)
            with c2:
                price_per_km = st.number_input("Tarif client (€ / km)", min_value=0.0, value=0.70, step=0.01)
            with c3:
                recovery_pct = st.number_input("Récupération appliquée (%)", min_value=0.0, max_value=200.0, value=recovery_default, step=5.0)
            run = st.form_submit_button("Calculer", use_container_width=True)
        if run or "fuel_normal_last" not in st.session_state:
            st.session_state["fuel_normal_last"] = {"km": km, "price_per_km": price_per_km, "recovery_pct": recovery_pct}
        vals = st.session_state["fuel_normal_last"]
        res = _compute_metrics(vals["km"], vals["price_per_km"], diesel_base, diesel_current, conso_100, vals["recovery_pct"])
        a, b, c, d = st.columns(4)
        a.metric("Carburant réel / km", _fmt_eur(res["cost_km_current"]))
        b.metric("Carburant trajet", _fmt_eur(res["fuel_total_current"], 2))
        c.metric("Surcharge / km", _fmt_eur(res["applied_surcharge_km"]))
        d.metric("Surcharge trajet", _fmt_eur(res["surcharge_total"], 2))
        e, f, g, h = st.columns(4)
        e.metric("Part carburant du tarif", _fmt_pct(res["price_share_pct"]))
        f.metric("% surcharge sur tarif", _fmt_pct(res["surcharge_pct"]))
        g.metric("Prix conseillé / km", _fmt_eur(res["recommended_price_km"]))
        h.metric("Total trajet avec surcharge", _fmt_eur(res["trip_total_with_surcharge"], 2))

    with tab_clients:
        st.markdown("### Règles client")
        st.caption("Le champ Match sert au calcul auto : il peut contenir un mot-clé du NOM, DEMANDEUR, BDC, GO ou DESTINATION.")
        st.dataframe(df_clients, use_container_width=True, hide_index=True, height=260)
        names = df_clients["Client"].tolist() if not df_clients.empty else []
        action = st.radio("Action", ["Ajouter", "Modifier / supprimer"], horizontal=True)
        selected_name = None
        selected_row = None
        if action == "Modifier / supprimer" and names:
            selected_name = st.selectbox("Client", names)
            selected_row = df_clients[df_clients["Client"] == selected_name].iloc[0]
        with st.form("fuel_client_form_pro", border=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                client_name = st.text_input("Nom client", value="" if selected_row is None else str(selected_row["Client"]))
            with c2:
                client_price = st.number_input("Prix / km", min_value=0.0, value=0.0 if selected_row is None else float(selected_row["Prix_km"]), step=0.01)
            with c3:
                client_recovery = st.number_input("Récupération %", min_value=0.0, max_value=200.0, value=recovery_default if selected_row is None else float(selected_row["Recuperation_pct"]), step=5.0)
            d1, d2 = st.columns(2)
            with d1:
                client_match = st.text_input("Texte de match auto", value="" if selected_row is None else str(selected_row.get("Match", "")))
            with d2:
                client_active = st.checkbox("Actif", value=True if selected_row is None else bool(selected_row["Actif"]))
            client_notes = st.text_input("Notes", value="" if selected_row is None else str(selected_row["Notes"]))
            save_client = st.form_submit_button("💾 Enregistrer", use_container_width=True)
            delete_client = st.form_submit_button("🗑️ Supprimer", use_container_width=True) if selected_row is not None else False
        if save_client:
            temp = df_clients.copy()
            if selected_name:
                temp = temp[temp["Client"] != selected_name]
            temp = pd.concat([
                temp,
                pd.DataFrame([{
                    "Client": str(client_name).strip(),
                    "Match": str(client_match).strip(),
                    "Prix_km": float(client_price),
                    "Recuperation_pct": float(client_recovery),
                    "Actif": bool(client_active),
                    "Notes": str(client_notes),
                }])
            ], ignore_index=True)
            _save_clients_to_db(temp)
            _reload_state()
            st.success("Règle client enregistrée.")
            st.rerun()
        if delete_client:
            temp = df_clients.copy()
            temp = temp[temp["Client"] != selected_name]
            _save_clients_to_db(temp)
            _reload_state()
            st.success("Règle client supprimée.")
            st.rerun()

    with tab_transfers:
        st.markdown("### Calcul sur un transfert précis")
        today = date.today()
        role = st.session_state.get("role")
        current_driver = str(st.session_state.get("chauffeur_code") or "").upper()

        p1, p2, p3 = st.columns(3)
        with p1:
            mode_period = st.selectbox("Période", ["Aujourd'hui", "Date précise", "Personnalisée"], index=0)
        with p2:
            start_date = st.date_input("Début", value=today)
        with p3:
            end_date = st.date_input("Fin", value=today)
        if mode_period == "Aujourd'hui":
            start_date = today
            end_date = today
        elif mode_period == "Date précise":
            end_date = start_date
        if end_date < start_date:
            st.error("La date de fin doit être >= à la date de début.")
            return

        df_rows = _load_planning_rows(start_date, end_date, only_active=True)
        if role != "admin" and current_driver in {"FA", "AD"}:
            df_rows = df_rows[df_rows["CH"].fillna("").astype(str).str.upper().str.contains(current_driver, regex=False)]

        if df_rows.empty:
            st.info("Aucun transfert trouvé sur la période.")
        else:
            df_view = df_rows[[c for c in ["DATE", "HEURE", "CH", "NOM", "DESIGNATION", "KM", "KM_EST", "SURCHARGE_CARBURANT"] if c in df_rows.columns]].copy()
            st.dataframe(df_view, use_container_width=True, hide_index=True, height=260)
            labels = [_transfer_label(r) for _, r in df_rows.iterrows()]
            chosen = st.selectbox("Transfert", labels)
            selected_idx = labels.index(chosen)
            row = df_rows.iloc[selected_idx]
            matched_rule = _find_matching_rule(row, df_clients)
            default_price = float(matched_rule["Prix_km"]) if matched_rule is not None else 0.70
            default_recovery = float(matched_rule["Recuperation_pct"]) if matched_rule is not None else recovery_default
            default_client_name = str(matched_rule["Client"]) if matched_rule is not None else ""
            km_effectif = _effective_km(row)
            if km_effectif <= 0:
                st.warning("Aucun KM exploitable sur ce transfert. Il faut idéalement remplir la colonne KM ; sinon le module utilise KM_EST s'il existe.")
            with st.form("fuel_selected_transfer_form", border=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    picked_client = st.text_input("Client / règle utilisée", value=default_client_name)
                with c2:
                    picked_price = st.number_input("Prix / km", min_value=0.0, value=default_price, step=0.01)
                with c3:
                    picked_recovery = st.number_input("Récupération %", min_value=0.0, max_value=200.0, value=default_recovery, step=5.0)
                km_override = st.number_input("KM retenus pour le calcul", min_value=0.0, value=float(km_effectif), step=1.0)
                save_db = st.form_submit_button("💾 Écrire en DB", use_container_width=True)
                save_excel = st.form_submit_button("📤 Écrire DB + XLSX", use_container_width=True)
            metrics = _compute_metrics(km_override, picked_price, diesel_base, diesel_current, conso_100, picked_recovery)
            a, b, c, d = st.columns(4)
            a.metric("Surcharge / km", _fmt_eur(metrics["applied_surcharge_km"]))
            b.metric("Surcharge trajet", _fmt_eur(metrics["surcharge_total"], 2))
            c.metric("% surcharge", _fmt_pct(metrics["surcharge_pct"]))
            d.metric("Prix conseillé / km", _fmt_eur(metrics["recommended_price_km"]))
            if save_db or save_excel:
                amount = round(metrics["surcharge_total"], 2)
                source = f"fuel_tab:{picked_client or 'manuel'}"
                _update_db_surcharge(int(row["id"]), amount, source)
                if save_excel:
                    ok, msg = _write_surcharge_to_excel(str(row.get("row_key", "")), amount, xlsx_header, xlsx_col_letter)
                    if ok:
                        st.success(f"Surcharge enregistrée en DB et dans Excel : {amount:.2f} €")
                    else:
                        st.warning(f"DB OK, Excel non mis à jour : {msg}")
                else:
                    st.success(f"Surcharge enregistrée en DB : {amount:.2f} €")

    with tab_auto:
        st.markdown("### Calcul automatique du jour / d'une période")
        c1, c2, c3 = st.columns(3)
        with c1:
            auto_mode = st.selectbox("Mode", ["Aujourd'hui", "Date précise", "Personnalisée"], index=0)
        with c2:
            auto_start = st.date_input("Début auto", value=date.today(), key="fuel_auto_start")
        with c3:
            auto_end = st.date_input("Fin auto", value=date.today(), key="fuel_auto_end")
        if auto_mode == "Aujourd'hui":
            auto_start = date.today()
            auto_end = date.today()
        elif auto_mode == "Date précise":
            auto_end = auto_start

        apply_excel = st.checkbox("Écrire aussi dans le XLSX Dropbox", value=True)
        if st.button("⚡ Calculer automatiquement", use_container_width=True):
            df_rows = _load_planning_rows(auto_start, auto_end, only_active=True)
            if df_rows.empty:
                st.info("Aucun transfert à traiter.")
            else:
                processed = 0
                skipped_no_rule = 0
                skipped_no_km = 0
                excel_errors = []
                preview_rows = []
                for _, row in df_rows.iterrows():
                    rule = _find_matching_rule(row, df_clients)
                    km_val = _effective_km(row)
                    if rule is None:
                        skipped_no_rule += 1
                        continue
                    if km_val <= 0:
                        skipped_no_km += 1
                        continue
                    m = _compute_metrics(km_val, float(rule["Prix_km"]), diesel_base, diesel_current, conso_100, float(rule["Recuperation_pct"]))
                    amount = round(m["surcharge_total"], 2)
                    _update_db_surcharge(int(row["id"]), amount, f"auto:{rule['Client']}")
                    if apply_excel:
                        ok, msg = _write_surcharge_to_excel(str(row.get("row_key", "")), amount, xlsx_header, xlsx_col_letter)
                        if not ok:
                            excel_errors.append(f"#{int(row['id'])}: {msg}")
                    preview_rows.append({
                        "DATE": row.get("DATE", ""),
                        "HEURE": row.get("HEURE", ""),
                        "CH": row.get("CH", ""),
                        "NOM": row.get("NOM", ""),
                        "KM": km_val,
                        "Règle": rule["Client"],
                        "Surcharge": amount,
                    })
                    processed += 1
                if preview_rows:
                    st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True, height=260)
                st.success(f"Calcul terminé : {processed} transfert(s) traités.")
                if skipped_no_rule:
                    st.warning(f"{skipped_no_rule} transfert(s) ignorés : aucune règle client trouvée.")
                if skipped_no_km:
                    st.warning(f"{skipped_no_km} transfert(s) ignorés : aucun KM exploitable.")
                if excel_errors:
                    st.warning("Certains transferts ont été écrits en DB mais pas dans le XLSX : " + " | ".join(excel_errors[:5]))
