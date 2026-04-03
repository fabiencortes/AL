# sync_core.py
"""
SYNC DROPBOX -> planning_cache (airportslines.db)
VERSION STANDALONE (hors Streamlit)
"""

import sqlite3
from datetime import datetime
from io import BytesIO

import requests
import pandas as pd
from openpyxl import load_workbook


# =====================================================
# CONFIG
# =====================================================

DB_PATH = "airportslines.db"
DROPBOX_FILE_PATH = "/Goldenlines/Planning 2026.xlsx"

# 🔐 clés (test uniquement)
DROPBOX_APP_KEY = "45g5h1k0cti3khz"
DROPBOX_APP_SECRET = "38yfl1yqtm40bc0"
DROPBOX_REFRESH_TOKEN = "3mq6k5EhfK8AAAAAAAAAAXQuWNymJzsIbyZjukwd41Ek6YFM2AsDWMIXXCkHV3qr"


# =====================================================
# LOG
# =====================================================

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


# =====================================================
# DROPBOX
# =====================================================

def get_dropbox_access_token():
    r = requests.post(
        "https://api.dropbox.com/oauth2/token",
        data={
            "grant_type": "refresh_token",
            "refresh_token": DROPBOX_REFRESH_TOKEN,
            "client_id": DROPBOX_APP_KEY,
            "client_secret": DROPBOX_APP_SECRET,
        },
        timeout=20,
    )
    r.raise_for_status()
    return r.json()["access_token"]


def download_excel_from_dropbox():
    token = get_dropbox_access_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Dropbox-API-Arg": f'{{"path": "{DROPBOX_FILE_PATH}"}}',
        "Content-Type": "application/octet-stream",
    }
    r = requests.post(
        "https://content.dropboxapi.com/2/files/download",
        headers=headers,
        timeout=60,
    )
    r.raise_for_status()
    return r.content


# =====================================================
# SQLITE
# =====================================================

def get_connection():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    return conn


# =====================================================
# SYNCHRO CORE
# =====================================================

def sync_planning_cache_from_dropbox():
    log("🚀 START sync_planning_cache_from_dropbox")

    # 1️⃣ Télécharger Excel
    content = download_excel_from_dropbox()
    log("📥 Excel téléchargé depuis Dropbox")

    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb["Feuil1"]

    headers = [c.value for c in ws[2]]
    rows = []

    for r in ws.iter_rows(min_row=3, values_only=True):
        if not any(r):
            continue
        rows.append(dict(zip(headers, r)))

    df = pd.DataFrame(rows)
    log(f"📊 {len(df)} lignes lues depuis Excel")

    if df.empty:
        log("⚠️ Excel vide → arrêt")
        return 0

    # 2️⃣ Connexion DB
    conn = get_connection()
    cur = conn.cursor()
    log("🔌 Connexion DB ouverte")

    # 3️⃣ RESET CACHE
    cur.execute("DELETE FROM planning_cache")
    log("🧹 planning_cache vidé")

    # 4️⃣ INSERT
    cols = [
        "DATE", "HEURE", "CH", "NOM", "ADRESSE", "CP", "LOCALITE",
        "DESIGNATION", "GO", "GROUPAGE", "PARTAGE", "VOL",
        "PAX", "PAIEMENT", "REMARQUE"
    ]

    inserted = 0

    for _, row in df.iterrows():
        values = [str(row.get(c, "") or "") for c in cols]

        cur.execute(
            f"""
            INSERT INTO planning_cache (
                {", ".join(cols)}
            ) VALUES (
                {", ".join(["?"] * len(cols))}
            )
            """,
            values,
        )
        inserted += 1

    conn.commit()
    conn.close()

    log(f"✅ DB mise à jour — {inserted} lignes insérées")
    log("🏁 END sync_planning_cache_from_dropbox")

    return inserted


# =====================================================
# MAIN
# =====================================================

if __name__ == "__main__":
    sync_planning_cache_from_dropbox()
