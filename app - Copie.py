import io
import urllib.parse
import ftplib
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st
from streamlit import column_config
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm

# =========================
#   CONFIG FTP
# =========================
FTP_HOST = "ftp.airports-linescom.webhosting.be"
FTP_USER = "info@airports-linescom"
FTP_PASSWORD = "A1rp0rts-L1nes"  # <-- tu peux changer ici si besoin
FTP_FILE_PATH = "/www/wp-content/uploads/2025/11/Planning-2025.xlsx"

# =========================
#   LISTE CHAUFFEURS (codes CH)
# =========================
CH_CODES = [
    "AU", "FA", "GD", "GG", "LL", "MA", "O", "RK", "RO", "SW", "NP", "DO", "OM",
    "AD", "CB", "CF", "CM", "EM", "GE", "HM", "JF", "KM", "LILLO", "MF", "WS", "FA1"
]


# =========================
#   FONCTIONS FTP
# =========================

def ftp_download_xlsx() -> bytes:
    """Télécharge le fichier XLSX depuis le FTP et renvoie les bytes."""
    bio = io.BytesIO()
    with ftplib.FTP(FTP_HOST, timeout=20) as ftp:
        ftp.login(FTP_USER, FTP_PASSWORD)
        ftp.retrbinary(f"RETR {FTP_FILE_PATH}", bio.write)
    return bio.getvalue()


def ftp_upload_xlsx(xlsx_bytes: bytes):
    """Envoie le fichier XLSX sur le FTP (remplace l'existant) + petit backup."""
    bio = io.BytesIO(xlsx_bytes)
    bio.seek(0)
    with ftplib.FTP(FTP_HOST, timeout=20) as ftp:
        ftp.login(FTP_USER, FTP_PASSWORD)

        # Backup sur le serveur FTP
        try:
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup_path = FTP_FILE_PATH.replace(".xlsx", f"-backup-{timestamp}.xlsx")
            ftp.storbinary(f"STOR {backup_path}", io.BytesIO(xlsx_bytes))
        except Exception:
            pass

        bio.seek(0)
        ftp.storbinary(f"STOR {FTP_FILE_PATH}", bio)


# =========================
#   ÉCRITURE DF -> WORKBOOK
# =========================

def write_df_to_sheet(ws, df: pd.DataFrame):
    """
    Écrit un DataFrame dans une feuille openpyxl, en gardant les styles :
      - ligne 1 = en-têtes existants
      - dès la ligne 2 = données du DF
    """
    df = df.copy()
    for col in ["_SELECT"]:
        if col in df.columns:
            df = df.drop(columns=[col])

    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col_index_by_name = {}
    for idx, name in enumerate(header_row, start=1):
        if name and name in df.columns:
            col_index_by_name[name] = idx

    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        for _, col_idx in col_index_by_name.items():
            ws.cell(row=row_idx, column=col_idx).value = None

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for col_name, col_idx in col_index_by_name.items():
            value = row.get(col_name, None)
            if pd.isna(value):
                value = None
            ws.cell(row=i, column=col_idx).value = value


def build_new_workbook_bytes(original_bytes: bytes,
                             df_feuil1: pd.DataFrame,
                             df_feuil2: pd.DataFrame,
                             df_feuil3: pd.DataFrame) -> bytes:
    """Réécrit seulement les valeurs de Feuil1/2/3 (styles conservés)."""
    bio = io.BytesIO(original_bytes)
    wb = load_workbook(bio, data_only=False)

    if "Feuil1" in wb.sheetnames:
        ws1 = wb["Feuil1"]
        write_df_to_sheet(ws1, df_feuil1)

    if "Feuil2" in wb.sheetnames and df_feuil2 is not None and not df_feuil2.empty:
        ws2 = wb["Feuil2"]
        write_df_to_sheet(ws2, df_feuil2)

    if "Feuil3" in wb.sheetnames and df_feuil3 is not None and not df_feuil3.empty:
        ws3 = wb["Feuil3"]
        write_df_to_sheet(ws3, df_feuil3)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
#   HELPERS GÉNÉRAUX
# =========================

def guess_client_column(df: pd.DataFrame):
    cols = list(df.columns)
    if "NOM" in cols:
        return "NOM"
    if "CLIENT" in cols:
        return "CLIENT"
    for c in cols:
        if "CLIENT" in c.upper():
            return c
    for c in cols:
        if "NOM" in c.upper():
            return c
    return None


def normalize_time_string(val: str) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""

    s = s.replace("H", "h").replace("h", ":").replace(" ", "")

    if s.isdigit():
        if len(s) <= 2:
            try:
                h = int(s)
            except ValueError:
                return val
            m = 0
        else:
            try:
                h = int(s[:-2])
                m = int(s[-2:])
            except ValueError:
                return val
        if 0 <= h <= 23 and 0 <= m <= 59:
            return f"{h:02d}:{m:02d}"
        return val

    if ":" in s:
        parts = s.split(":")
        if len(parts) != 2:
            return val
        try:
            h = int(parts[0])
            m = int(parts[1])
        except ValueError:
            return val
        if 0 <= h <= 23 and 0 <= m <= 59:
            return f"{h:02d}:{m:02d}"
        return val

    return val


def detect_chauffeur_conflict(df: pd.DataFrame, new_row: dict, current_idx=None) -> bool:
    if "CH" not in df.columns or "DATE" not in df.columns or "HEURE" not in df.columns:
        return False

    ch = str(new_row.get("CH", "")).strip()
    date_val = new_row.get("DATE", None)
    heure_val = str(new_row.get("HEURE", "")).strip()

    if not ch or not date_val or not heure_val:
        return False

    mask = (df["CH"].astype(str).str.strip() == ch) & \
           (df["DATE"] == date_val) & \
           (df["HEURE"].astype(str).str.strip() == heure_val)

    if current_idx is not None:
        mask &= (df.index != current_idx)

    conflicts = df.index[mask].tolist()
    return len(conflicts) > 0


def extract_chauffeurs_from_cell(value) -> list:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []

    text = str(value).upper()
    detected = []
    for code in CH_CODES:
        if code in text:
            detected.append(code)
    return list(dict.fromkeys(detected))


def sort_heure_for_display(s: str):
    if not isinstance(s, str):
        s = str(s) if s is not None else ""
    s = s.strip()
    if not s:
        return (99, 99)
    s_norm = normalize_time_string(s)
    try:
        h, m = s_norm.split(":")
        h = int(h)
        m = int(m)
        return (h, m)
    except Exception:
        return (99, 99)


# =========================
#   HELPERS TEL / WHATSAPP / MAILTO / PDF
# =========================

def clean_phone(phone: str) -> str:
    if phone is None:
        return ""
    return "".join(ch for ch in str(phone) if ch.isdigit())


def phone_to_whatsapp_number(phone: str) -> str:
    digits = clean_phone(phone)
    if not digits:
        return ""
    if digits.startswith("0"):
        return "32" + digits[1:]
    return digits


def build_whatsapp_link(phone: str, message: str) -> str:
    num = phone_to_whatsapp_number(phone)
    if not num:
        return "#"
    text = urllib.parse.quote(message)
    return f"https://wa.me/{num}?text={text}"


def build_mailto_link(to_email: str, subject: str, body: str) -> str:
    if not to_email:
        return "#"
    subj_encoded = urllib.parse.quote(subject)
    body_encoded = urllib.parse.quote(body)
    return f"mailto:{to_email}?subject={subj_encoded}&body={body_encoded}"


def create_chauffeur_pdf(df_ch: pd.DataFrame, ch_selected: str, day_selected: date) -> bytes:
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    y = height - 2 * cm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2 * cm, y, f"Feuille chauffeur - {ch_selected} - {day_selected.strftime('%d/%m/%Y')}")
    y -= 1 * cm
    c.setFont("Helvetica", 10)

    for _, row in df_ch.iterrows():
        if y < 3 * cm:
            c.showPage()
            y = height - 2 * cm
            c.setFont("Helvetica-Bold", 14)
            c.drawString(2 * cm, y, f"Feuille chauffeur - {ch_selected} - {day_selected.strftime('%d/%m/%Y')}")
            y -= 1 * cm
            c.setFont("Helvetica", 10)

        heure = normalize_time_string(row.get("HEURE", ""))

        designation = str(row.get("DESIGNATION", "") or "")
        route_txt = ""
        for cand in ["Unnamed: 8", "DESIGNATION"]:
            if cand in df_ch.columns:
                val = row.get(cand, "")
                if pd.notna(val) and str(val).strip():
                    route_txt = str(val).strip()
                    break

        if route_txt and designation and designation not in route_txt:
            dest_full = f"{route_txt} ({designation})"
        elif route_txt:
            dest_full = route_txt
        elif designation:
            dest_full = designation
        else:
            dest_full = "Navette"

        nom = str(row.get("NOM", "") or "")
        adresse = str(row.get("ADRESSE", "") or "")
        cp = str(row.get("CP", "") or "")
        localite = str(row.get("Localité", "") or row.get("LOCALITE", "") or "")
        pax = str(row.get("PAX", "") or "")
        paiement = str(row.get("PAIEMENT", "") or "")
        caisse = str(row.get("Caisse", "") or "")
        num_vol = str(row.get("N° Vol", "") or "")
        origine = str(row.get("Origine", "") or "")
        decol = str(row.get("Décollage", "") or "")
        h_south = str(row.get("H South", "") or "")

        adr_full = " ".join(x for x in [adresse, cp, localite] if x)

        ligne1 = f"{heure}  -  {dest_full}"
        if nom:
            ligne1 += f" - {nom}"
        c.drawString(2 * cm, y, ligne1)
        y -= 0.5 * cm

        if adr_full:
            c.drawString(2 * cm, y, adr_full)
            y -= 0.5 * cm

        infos_vol = []
        if num_vol:
            infos_vol.append(f"Vol {num_vol}")
        if origine:
            infos_vol.append(f"Origine {origine}")
        if decol:
            infos_vol.append(f"Décollage {decol}")
        if h_south:
            infos_vol.append(f"H SO {h_south}")
        if infos_vol:
            c.drawString(2 * cm, y, " | ".join(infos_vol))
            y -= 0.5 * cm

        infos_pay = []
        if pax:
            infos_pay.append(f"PAX {pax}")
        if paiement:
            infos_pay.append(f"Paiement {paiement}")
        if caisse:
            infos_pay.append(f"Caisse {caisse}€")
        if infos_pay:
            c.drawString(2 * cm, y, " | ".join(infos_pay))
            y -= 0.5 * cm

        y -= 0.5 * cm

    c.save()
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


def build_chauffeur_day_message(df_ch: pd.DataFrame, ch_selected: str, day_selected: date) -> str:
    lines = []
    header = f"🚖 Planning du {day_selected.strftime('%d/%m/%Y')} — Chauffeur : {ch_selected}"
    lines.append(header)
    lines.append("")

    for _, row in df_ch.iterrows():
        heure = normalize_time_string(row.get("HEURE", ""))
        nom = str(row.get("NOM", "") or "")

        designation = str(row.get("DESIGNATION", "") or "")
        route_txt = ""
        for cand in ["Unnamed: 8", "DESIGNATION"]:
            if cand in df_ch.columns:
                val = row.get(cand, "")
                if pd.notna(val) and str(val).strip():
                    route_txt = str(val).strip()
                    break

        if route_txt and designation and designation not in route_txt:
            dest_full = f"{route_txt} ({designation})"
        elif route_txt:
            dest_full = route_txt
        elif designation:
            dest_full = designation
        else:
            dest_full = "Navette"

        adresse = str(row.get("ADRESSE", "") or "")
        cp = str(row.get("CP", "") or "")
        localite = str(row.get("Localité", "") or row.get("LOCALITE", "") or "")
        adr_full = " ".join(x for x in [adresse, cp, localite] if x)

        pax = str(row.get("PAX", "") or "")
        paiement = str(row.get("PAIEMENT", "") or "")
        caisse = str(row.get("Caisse", "") or "")

        line1 = f"➡ {heure or '??:??'} — {dest_full}"
        if nom:
            line1 += f" — {nom}"
        lines.append(line1)
        if adr_full:
            lines.append(f"   📍 {adr_full}")
        extra = []
        if pax:
            extra.append(f"{pax} pax")
        if paiement:
            extra.append(f"Paiement: {paiement}")
        if caisse:
            extra.append(f"Caisse: {caisse} €")
        if extra:
            lines.append("   " + " — ".join(extra))
        lines.append("")

    return "\n".join(lines).strip()


def build_chauffeur_future_message(df: pd.DataFrame, ch_selected: str, from_date: date) -> str:
    lines = []
    lines.append(f"🚖 Planning à partir du {from_date.strftime('%d/%m/%Y')} — Chauffeur : {ch_selected}")
    lines.append("")

    if "DATE" in df.columns:
        try:
            df_date = df.copy()
            if not pd.api.types.is_datetime64_any_dtype(df_date["DATE"]):
                df_date["DATE"] = pd.to_datetime(df_date["DATE"], errors="coerce").dt.date
        except Exception:
            df_date = df.copy()
    else:
        df_date = df.copy()

    if "DATE" in df_date.columns:
        df_date = df_date[df_date["DATE"] >= from_date].copy()

    if df_date.empty:
        lines.append("Aucune navette planifiée.")
        return "\n".join(lines)

    def row_has_ch(row):
        ch_list = extract_chauffeurs_from_cell(row.get("CH", ""))
        return ch_selected in ch_list

    df_date["__HAS_CH__"] = df_date.apply(row_has_ch, axis=1)
    df_ch = df_date[df_date["__HAS_CH__"] == True].copy()
    df_ch = df_ch.drop(columns=["__HAS_CH__"])

    if df_ch.empty:
        lines.append("Aucune navette planifiée pour ce chauffeur.")
        return "\n".join(lines)

    if "DATE" in df_ch.columns:
        try:
            df_ch["DATE"] = pd.to_datetime(df_ch["DATE"], errors="coerce").dt.date
        except Exception:
            pass
    if "HEURE" in df_ch.columns:
        df_ch["__HEURE_SORT__"] = df_ch["HEURE"].apply(sort_heure_for_display)
        sort_cols = []
        if "DATE" in df_ch.columns:
            sort_cols.append("DATE")
        sort_cols.append("__HEURE_SORT__")
        df_ch = df_ch.sort_values(sort_cols).drop(columns=["__HEURE_SORT__"])

    if "DATE" in df_ch.columns:
        grouped = df_ch.groupby("DATE")
    else:
        grouped = [(None, df_ch)]

    for d, subdf in grouped:
        if d is not None:
            lines.append(f"📆 {d.strftime('%d/%m/%Y')}")
        else:
            lines.append("📆 Jour non défini")

        for _, row in subdf.iterrows():
            heure = normalize_time_string(row.get("HEURE", ""))
            nom = str(row.get("NOM", "") or "")

            designation = str(row.get("DESIGNATION", "") or "")
            route_txt = ""
            for cand in ["Unnamed: 8", "DESIGNATION"]:
                if cand in df_ch.columns:
                    val = row.get(cand, "")
                    if pd.notna(val) and str(val).strip():
                        route_txt = str(val).strip()
                        break

            if route_txt and designation and designation not in route_txt:
                dest_full = f"{route_txt} ({designation})"
            elif route_txt:
                dest_full = route_txt
            elif designation:
                dest_full = designation
            else:
                dest_full = "Navette"

            adresse = str(row.get("ADRESSE", "") or "")
            cp = str(row.get("CP", "") or "")
            localite = str(row.get("Localité", "") or row.get("LOCALITE", "") or "")
            adr_full = " ".join(x for x in [adresse, cp, localite] if x)

            pax = str(row.get("PAX", "") or "")
            paiement = str(row.get("PAIEMENT", "") or "")
            caisse = str(row.get("Caisse", "") or "")

            line1 = f"  ➡ {heure or '??:??'} — {dest_full}"
            if nom:
                line1 += f" — {nom}"
            lines.append(line1)
            if adr_full:
                lines.append(f"     📍 {adr_full}")
            extra = []
            if pax:
                extra.append(f"{pax} pax")
            if paiement:
                extra.append(f"Paiement: {paiement}")
            if caisse:
                extra.append(f"Caisse: {caisse} €")
            if extra:
                lines.append("     " + " — ".join(extra))
            lines.append("")

        lines.append("")

    return "\n".join(lines).strip()


# =========================
#   APP STREAMLIT
# =========================

st.set_page_config(
    page_title="Airports-Lines — Planning chauffeurs",
    layout="wide",
)


def init_session_state():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "xlsx_bytes" not in st.session_state:
        st.session_state.xlsx_bytes = None
    if "df_feuil1" not in st.session_state:
        st.session_state.df_feuil1 = None
    if "df_feuil2" not in st.session_state:
        st.session_state.df_feuil2 = None
    if "df_feuil3" not in st.session_state:
        st.session_state.df_feuil3 = None


def load_data_from_ftp():
    try:
        with st.spinner("Chargement du planning depuis le serveur FTP…"):
            xbytes = ftp_download_xlsx()

            bio1 = io.BytesIO(xbytes)
            df1 = pd.read_excel(bio1, sheet_name="Feuil1", dtype=str, engine="openpyxl")
            try:
                bio2 = io.BytesIO(xbytes)
                df2 = pd.read_excel(bio2, sheet_name="Feuil2", dtype=str, engine="openpyxl")
            except Exception:
                df2 = pd.DataFrame()
            try:
                bio3 = io.BytesIO(xbytes)
                df3 = pd.read_excel(bio3, sheet_name="Feuil3", dtype=str, engine="openpyxl")
            except Exception:
                df3 = pd.DataFrame()

            # On ne touche qu'à DATE (reste = exactement ce qui vient d'Excel)
            if "DATE" in df1.columns:
                try:
                    df1["DATE"] = pd.to_datetime(df1["DATE"], errors="coerce").dt.date
                except Exception:
                    pass

            if "GROUPAGE" not in df1.columns:
                df1["GROUPAGE"] = False
            if "PARTAGE" not in df1.columns:
                df1["PARTAGE"] = False

            if df2 is not None and not df2.empty and df2.shape[1] >= 3:
                third_col_name = df2.columns[2]
                if (not isinstance(third_col_name, str)) or str(third_col_name).startswith("Unnamed"):
                    df2 = df2.rename(columns={third_col_name: "TEL_CH"})

            st.session_state.xlsx_bytes = xbytes
            st.session_state.df_feuil1 = df1.reset_index(drop=True)
            st.session_state.df_feuil2 = df2.reset_index(drop=True) if not df2.empty else df2
            st.session_state.df_feuil3 = df3.reset_index(drop=True) if not df3.empty else df3

    except Exception as e:
        st.session_state.xlsx_bytes = None
        st.session_state.df_feuil1 = None
        st.session_state.df_feuil2 = None
        st.session_state.df_feuil3 = None
        st.error(f"❌ Erreur lors du chargement depuis le FTP : {e}")


def login_screen():
    st.title("🚐 Airports-Lines — Planning chauffeurs")
    st.subheader("Connexion administrateur")
    st.write("Pour l'instant : **login = admin / mot de passe = admin**")

    col1, col2 = st.columns(2)
    with col1:
        login = st.text_input("Login", value="", key="login_name")
    with col2:
        pwd = st.text_input("Mot de passe", value="", type="password", key="login_pass")

    if st.button("Se connecter"):
        if login == "admin" and pwd == "admin":
            st.session_state.logged_in = True
            load_data_from_ftp()
            st.success("Connecté ✅")
            st.rerun()
        else:
            st.error("Identifiants incorrects.")


def render_top_bar():
    st.markdown("### 🚐 Airports-Lines — Gestion du planning (serveur FTP)")

    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        if st.button("🔄 Recharger depuis le FTP"):
            load_data_from_ftp()
    with col2:
        if st.button("💾 Enregistrer vers le FTP"):
            if st.session_state.xlsx_bytes is None:
                st.error("Aucun fichier chargé.")
            else:
                try:
                    new_bytes = build_new_workbook_bytes(
                        st.session_state.xlsx_bytes,
                        st.session_state.df_feuil1,
                        st.session_state.df_feuil2,
                        st.session_state.df_feuil3,
                    )
                    with st.spinner("Envoi du fichier modifié vers le serveur FTP…"):
                        ftp_upload_xlsx(new_bytes)
                    st.session_state.xlsx_bytes = new_bytes
                    st.success("✅ Planning enregistré sur le serveur FTP.")
                except Exception as e:
                    st.error(f"Erreur lors de la sauvegarde sur FTP : {e}")
    with col3:
        st.info("Mode **manuel** : rien n'est rechargé ni sauvegardé sur le FTP tant que tu n'appuies pas sur les boutons.")


# =========================
#   ONGLET 📅 PLANNING
# =========================

def render_tab_planning():
    if st.session_state.df_feuil1 is None:
        st.warning("Aucun planning chargé. Utilise le bouton 🔄 Recharger depuis le FTP.")
        return

    st.subheader("📅 Planning — vue agenda")

    df = st.session_state.df_feuil1.copy()

    if "GROUPAGE" not in df.columns:
        df["GROUPAGE"] = False
    if "PARTAGE" not in df.columns:
        df["PARTAGE"] = False

    today = date.today()

    colf1, colf2, colf3, colf4 = st.columns([1, 1, 1, 2])
    with colf1:
        d_deb = st.date_input("Date de début", value=today)
    with colf2:
        d_fin = st.date_input("Date de fin", value=today)
    with colf3:
        ch_value = None
        if "CH" in df.columns:
            options = sorted(set(str(x).strip() for x in df["CH"].dropna() if str(x).strip()))
            ch_value = st.selectbox("Chauffeur (CH)", ["(Tous)"] + options)
            if ch_value == "(Tous)":
                ch_value = None
    with colf4:
        col4a, col4b = st.columns([2, 1])
        with col4a:
            search = st.text_input("Recherche (client, désignation, vol, remarque…)", "")
        with col4b:
            tri_par_chauffeur = st.checkbox("Trier par chauffeur", value=False, key="tri_ch_planning")

    # Normalisation des dates + filtre période, SANS apply / SANS keep_in_interval
    if "DATE" in df.columns:
        try:
            df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce")
        except Exception:
            pass

        mask_date = df["DATE"].notna() & df["DATE"].dt.date.between(d_deb, d_fin)
        df = df[mask_date].copy().reset_index(drop=True)

    # Filtre chauffeur
    if ch_value is not None and "CH" in df.columns:
        mask_ch = df["CH"].astype(str).str.strip() == str(ch_value).strip()
        df = df[mask_ch].copy().reset_index(drop=True)

    # Filtre texte global
    if search.strip():
        s_low = search.lower()
        mask = pd.Series(False, index=df.index)
        for col in df.columns:
            mask = mask | df[col].astype(str).str.lower().str.contains(s_low, na=False)
        df = df[mask].copy().reset_index(drop=True)

    if df.empty:
        st.warning("Aucune course pour les filtres sélectionnés.")
        return

    # Tri
    sort_cols = []
    if tri_par_chauffeur and "CH" in df.columns:
        sort_cols.append("CH")
    if "DATE" in df.columns:
        sort_cols.append("DATE")
    if "HEURE" in df.columns:
        df["HEURE_SORT"] = df["HEURE"].apply(sort_heure_for_display)
        sort_cols.append("HEURE_SORT")

    if sort_cols:
        df = df.sort_values(sort_cols).reset_index(drop=True)
        if "HEURE_SORT" in df.columns:
            df = df.drop(columns=["HEURE_SORT"])

    # Colonne TYPE (groupage / partagé)
    def type_from_flags(row):
        try:
            g = str(row.get("GROUPAGE", "")).lower() in ["true", "1", "oui", "x"]
            p = str(row.get("PARTAGE", "")).lower() in ["true", "1", "oui", "x"]
        except Exception:
            g = False
            p = False
        if g:
            return "🔶 Groupage"
        if p:
            return "🟨 Partagée"
        return ""

    df_display = df.copy()
    df_display["TYPE"] = df_display.apply(type_from_flags, axis=1)

    # Regroupement par DATE pour l'affichage
    if "DATE" in df_display.columns:
        grouped = df_display.groupby("DATE")
    else:
        grouped = [(None, df_display)]

    for d, subdf in grouped:
        d_display = None
        if isinstance(d, (pd.Timestamp, datetime)):
            d_display = d.date()
        elif isinstance(d, date):
            d_display = d

        if d_display is not None:
            st.markdown(f"#### 📆 {d_display.strftime('%d/%m/%Y')}")
        else:
            st.markdown("#### 📆 Jour non défini")

        st.dataframe(subdf, use_container_width=True)


# =========================
#   TABLEAU FEUIL1
# =========================

def render_tab_table():
    if st.session_state.df_feuil1 is None:
        st.warning("Aucun planning chargé.")
        return

    st.subheader("📊 Tableau Feuil1 — filtre par date, coche une ligne pour la fiche détaillée")

    df = st.session_state.df_feuil1.copy().reset_index(drop=True)

    if "GROUPAGE" not in df.columns:
        df["GROUPAGE"] = False
    if "PARTAGE" not in df.columns:
        df["PARTAGE"] = False

    if "DATE" in df.columns:
        try:
            df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce")
        except Exception:
            pass

        col_fdate1, col_fdate2 = st.columns([1, 1])
        with col_fdate1:
            filter_enabled = st.checkbox("Activer le filtre par date", value=True, key="table_filter_enabled")
        with col_fdate2:
            min_date = df["DATE"].dropna().min()
            if isinstance(min_date, pd.Timestamp):
                min_date_val = min_date.date()
            else:
                min_date_val = date.today()
            start_date = st.date_input(
                "Afficher les navettes à partir du :",
                value=min_date_val,
                key="table_start_date",
            )

        if filter_enabled:
            def is_on_or_after_start(d):
                if pd.isna(d):
                    return False
                if isinstance(d, pd.Timestamp):
                    d_date = d.date()
                elif isinstance(d, datetime):
                    d_date = d.date()
                elif isinstance(d, date):
                    d_date = d
                else:
                    return False
                return d_date >= start_date

            mask = df["DATE"].apply(is_on_or_after_start)
            df = df[mask].reset_index(drop=True)

        if df.empty:
            st.warning("Aucune navette avec ces paramètres.")
            return

        df["DATE"] = df["DATE"].dt.date

    if st.button("⬇️ Trier par DATE + HEURE (Feuil1)"):
        df_sort = df.copy()
        if "DATE" in df_sort.columns:
            df_sort["DATE_SORT"] = pd.to_datetime(df_sort["DATE"], errors="coerce")
        else:
            df_sort["DATE_SORT"] = pd.NaT
        if "HEURE" in df_sort.columns:
            df_sort["HEURE_SORT"] = df_sort["HEURE"].apply(sort_heure_for_display)
        else:
            df_sort["HEURE_SORT"] = (99, 99)
        df_sort = df_sort.sort_values(["DATE_SORT", "HEURE_SORT"]).reset_index(drop=True)
        df_sort = df_sort.drop(columns=["DATE_SORT", "HEURE_SORT"])
        st.session_state.df_feuil1 = df_sort.copy()
        st.success("Tri appliqué. Pour l'enregistrer sur le FTP, clique sur 💾 Enregistrer vers le FTP en haut.")
        df = df_sort.copy()

    df.insert(0, "_SELECT", False)

    df2 = st.session_state.df_feuil2
    ch_options_grid = None
    if df2 is not None and not df2.empty and "CH" in df2.columns:
        ch_options_grid = sorted(set(str(x).strip() for x in df2["CH"].dropna() if str(x).strip()))
    elif "CH" in df.columns:
        ch_options_grid = sorted(set(str(x).strip() for x in df["CH"].dropna() if str(x).strip()))

    col_conf = {
        "_SELECT": column_config.CheckboxColumn(
            "✔",
            help="Coche ici pour sélectionner la ligne.",
            default=False,
        )
    }
    if "CH" in df.columns and ch_options_grid:
        col_conf["CH"] = column_config.SelectboxColumn(
            "CH",
            options=ch_options_grid,
            help="Choisis un chauffeur (tape l'initiale pour filtrer).",
        )
    if "GROUPAGE" in df.columns:
        col_conf["GROUPAGE"] = column_config.CheckboxColumn(
            "Groupage",
            help="Cocher si cette navette fait partie d'un groupage.",
            default=False,
        )
    if "PARTAGE" in df.columns:
        col_conf["PARTAGE"] = column_config.CheckboxColumn(
            "Navette partagée",
            help="Cocher si la navette est partagée.",
            default=False,
        )

    edited = st.data_editor(
        df,
        use_container_width=True,
        num_rows="dynamic",
        key="editor_feuil1",
        column_config=col_conf,
    )

    edited = edited.reset_index(drop=True)
    selected_indices = edited.index[edited["_SELECT"] == True].tolist()
    if selected_indices:
        idx = selected_indices[-1]
    else:
        idx = 0

    df_clean = edited.drop(columns=["_SELECT"])
    st.session_state.df_feuil1 = df_clean

    if df_clean.empty:
        st.info("Le tableau est vide.")
        return

    base_row = df_clean.iloc[idx].to_dict()
    client_col = guess_client_column(df_clean)

    st.markdown(f"#### 🧳 Ligne sélectionnée : {idx + 1}")
    st.dataframe(df_clean.iloc[[idx]], use_container_width=True)

    groupage_initial = bool(base_row.get("GROUPAGE", False))
    st.markdown("### 👥 Groupage et client")

    col_g1, col_g2 = st.columns([1, 3])
    with col_g1:
        groupage_checked = st.checkbox(
            "Groupage", value=groupage_initial, key=f"groupage_flag_feuil1_{idx}"
        )

    if groupage_checked and client_col:
        mode = st.radio(
            "Client pour cette navette groupée",
            ["Client existant", "Nouveau client"],
            key=f"groupage_mode_feuil1_{idx}",
            horizontal=True,
        )
        all_clients = sorted(
            set(str(x).strip() for x in df_clean[client_col].dropna() if str(x).strip())
        )

        if mode == "Client existant":
            selected_client_for_groupage = st.selectbox(
                "Choisir un client existant",
                options=[""] + all_clients,
                key=f"groupage_existing_client_feuil1_{idx}",
            )
            if selected_client_for_groupage:
                client_rows = df_clean[
                    df_clean[client_col].astype(str).str.strip()
                    == selected_client_for_groupage.strip()
                ]
                if not client_rows.empty:
                    try:
                        tmp = client_rows.copy()
                        if "DATE" in tmp.columns:
                            tmp["DATE_TMP"] = pd.to_datetime(tmp["DATE"], errors="coerce")
                        else:
                            tmp["DATE_TMP"] = pd.NaT
                        if "HEURE" in tmp.columns:
                            tmp["HEURE_TMP"] = tmp["HEURE"]
                        else:
                            tmp["HEURE_TMP"] = ""
                        tmp = tmp.sort_values(["DATE_TMP", "HEURE_TMP"])
                        client_row = tmp.iloc[-1]
                    except Exception:
                        client_row = client_rows.iloc[-1]

                    client_info_cols = [
                        client_col,
                        "ADRESSE", "Adresse",
                        "CP",
                        "Localité", "LOCALITE", "LOCALITÉ",
                        "Tél", "TEL",
                        "Type Nav", "PAIEMENT", "Caisse",
                    ]
                    for c in client_info_cols:
                        if c in df_clean.columns and c in client_row.index:
                            base_row[c] = client_row[c]

                    st.info("Fiche client chargée comme modèle pour cette navette.")
                    st.write("Dernière navette de ce client :")
                    st.dataframe(client_row.to_frame().T, use_container_width=True)
        else:
            new_client_name = st.text_input(
                "Nom du nouveau client pour ce groupage",
                value=str(base_row.get(client_col, "")) if base_row.get(client_col) else "",
                key=f"groupage_new_client_feuil1_{idx}",
            )
            if new_client_name:
                base_row[client_col] = new_client_name

    st.markdown("### 📝 Fiche détaillée de la navette sélectionnée")

    new_row = {}
    cols_left, cols_right = st.columns(2)
    columns = list(df_clean.columns)

    df2 = st.session_state.df_feuil2
    ch_options_form = None
    if df2 is not None and not df2.empty and "CH" in df2.columns:
        ch_options_form = sorted(set(str(x).strip() for x in df2["CH"].dropna() if str(x).strip()))
    elif "CH" in df_clean.columns:
        ch_options_form = sorted(set(str(x).strip() for x in df_clean["CH"].dropna() if str(x).strip()))

    for i, col in enumerate(columns):
        if col == "GROUPAGE":
            new_row[col] = groupage_checked
            continue

        container = cols_left if i % 2 == 0 else cols_right
        val = base_row.get(col, "")

        widget_key_prefix = f"feuil1_form_{col}_{idx}"

        if col == "DATE":
            default_date = date.today()
            if isinstance(val, (date, datetime)):
                default_date = val if isinstance(val, date) else val.date()
            else:
                try:
                    ts = pd.to_datetime(val, errors="coerce")
                    if not pd.isna(ts):
                        default_date = ts.date()
                except Exception:
                    default_date = date.today()
            new_val = container.date_input("DATE", value=default_date, key=widget_key_prefix)
            new_row[col] = new_val
            continue

        if col == "PARTAGE":
            bool_default = False
            if isinstance(val, bool):
                bool_default = val
            else:
                try:
                    bool_default = str(val).lower() in ["true", "1", "oui", "x"]
                except Exception:
                    bool_default = False
            new_val = container.checkbox(
                "Navette partagée",
                value=bool_default,
                key=widget_key_prefix,
            )
            new_row[col] = new_val
            continue

        if col == "CH" and ch_options_form:
            current = str(val).strip() if val is not None else ""
            opts = [""] + ch_options_form
            try:
                default_index = opts.index(current)
            except ValueError:
                default_index = 0
            new_val = container.selectbox(
                "CH (chauffeur)",
                opts,
                index=default_index,
                key=widget_key_prefix,
            )
            new_row[col] = new_val
            continue

        if col == "HEURE":
            text_val = "" if val is None or pd.isna(val) else str(val)
            raw_val = container.text_input(col, value=text_val, key=widget_key_prefix)
            new_row[col] = normalize_time_string(raw_val)
            continue

        text_val = "" if val is None or pd.isna(val) else str(val)
        new_val = container.text_input(col, value=text_val, key=widget_key_prefix)
        new_row[col] = new_val

    new_row["GROUPAGE"] = groupage_checked

    st.markdown("#### 🚦 Actions sur cette fiche")

    c1, c2, c3, c4, c5 = st.columns(5)

    with c1:
        if st.button("✅ Mettre à jour la ligne sélectionnée"):
            df_new = df_clean.copy()
            if detect_chauffeur_conflict(df_new, new_row, current_idx=idx):
                st.warning("⚠ Conflit chauffeur sur DATE / HEURE, mise à jour quand même.")
            for k, v in new_row.items():
                df_new.at[idx, k] = v
            st.session_state.df_feuil1 = df_new.reset_index(drop=True)
            st.success("Ligne mise à jour.")
            st.rerun()

    with c2:
        if st.button("➕ Créer une nouvelle navette à partir de cette fiche"):
            df_new = df_clean.copy()
            if detect_chauffeur_conflict(df_new, new_row, current_idx=None):
                st.warning("⚠ Conflit chauffeur sur DATE / HEURE, navette créée quand même.")
            df_new = pd.concat([df_new, pd.DataFrame([new_row])], ignore_index=True)
            st.session_state.df_feuil1 = df_new.reset_index(drop=True)
            st.success("Nouvelle navette ajoutée au planning.")
            st.rerun()

    with c3:
        if st.button("🗑️ Supprimer la ligne sélectionnée"):
            df_new = df_clean.drop(df_clean.index[idx]).reset_index(drop=True)
            st.session_state.df_feuil1 = df_new
            st.success("Ligne supprimée.")
            st.rerun()

    with c4:
        if st.button("📆 Dupliquer pour demain"):
            df_new = df_clean.copy()
            clone = new_row.copy()
            dval = clone.get("DATE")
            if isinstance(dval, (date, datetime)):
                clone["DATE"] = dval + timedelta(days=1)
            df_new = pd.concat([df_new, pd.DataFrame([clone])], ignore_index=True)
            st.session_state.df_feuil1 = df_new.reset_index(drop=True)
            st.success("Navette dupliquée pour le lendemain.")
            st.rerun()

    with c5:
        new_date = st.date_input(
            "Date pour duplication personnalisée", value=date.today(), key=f"dup_custom_date_{idx}"
        )
        if st.button("📆 Dupliquer sur cette date"):
            df_new = df_clean.copy()
            clone = new_row.copy()
            clone["DATE"] = new_date
            df_new = pd.concat([df_new, pd.DataFrame([clone])], ignore_index=True)
            st.session_state.df_feuil1 = df_new.reset_index(drop=True)
            st.success(f"Navette dupliquée pour le {new_date}.")
            st.rerun()


# =========================
#   FEUIL2 / FEUIL3
# =========================

def render_tab_feuil2():
    st.subheader("👨‍✈️ Feuil2 — Chauffeurs")
    st.write("Ici tu gères la liste des chauffeurs (codes, noms, téléphones…).")

    df2 = st.session_state.df_feuil2
    if df2 is None or df2.empty:
        st.info("Aucune feuille Feuil2 trouvée dans le fichier.")
        return

    edited = st.data_editor(
        df2,
        use_container_width=True,
        num_rows="dynamic",
        key="editor_feuil2",
    )
    st.session_state.df_feuil2 = edited.reset_index(drop=True)


def render_tab_feuil3():
    st.subheader("📄 Feuil3 — Données annexes")
    st.write("Cette feuille est modifiable mais séparée du planning.")

    df3 = st.session_state.df_feuil3
    if df3 is None or df3.empty:
        st.info("Aucune feuille Feuil3 trouvée dans le fichier.")
        return

    edited = st.data_editor(
        df3,
        use_container_width=True,
        num_rows="dynamic",
        key="editor_feuil3",
    )
    st.session_state.df_feuil3 = edited.reset_index(drop=True)


# =========================
#   CLIENTS / HISTORIQUE
# =========================

def render_tab_clients():
    st.subheader("🔍 Clients — Historique & création rapide")

    df = st.session_state.df_feuil1
    if df is None:
        st.warning("Aucun planning chargé.")
        return

    if "GROUPAGE" not in df.columns:
        df["GROUPAGE"] = False
    if "PARTAGE" not in df.columns:
        df["PARTAGE"] = False

    client_col = guess_client_column(df)
    if client_col is None:
        st.info("Impossible de trouver une colonne 'client' (NOM / CLIENT...).")
        return

    st.write(f"Colonne client utilisée : **{client_col}**")

    col_search, col_info = st.columns([2, 1])
    with col_search:
        query = st.text_input("Nom du client (ou partie du nom)", "")
    with col_info:
        st.caption("Laisse vide pour créer une navette pour un nouveau client.")

    df_clients = pd.DataFrame()
    base_row = None
    orig_idx = None

    if query.strip():
        mask = df[client_col].astype(str).str.contains(query.strip(), case=False, na=False)
        df_clients = df[mask].copy()
        if df_clients.empty:
            st.warning("Aucun transfert trouvé pour ce client.")
        else:
            df_clients = df_clients.reset_index().rename(columns={"index": "_ORIG_INDEX_"})
            df_clients.insert(0, "_SELECT", False)

            st.markdown("**Transferts trouvés pour ce client (coche une ligne) :**")

            col_conf = {
                "_SELECT": column_config.CheckboxColumn(
                    "✔",
                    help="Coche ici pour sélectionner la navette comme modèle.",
                    default=False,
                )
            }

            edited_clients = st.data_editor(
                df_clients,
                use_container_width=True,
                num_rows="fixed",
                key="clients_editor",
                column_config=col_conf,
            )

            edited_clients = edited_clients.reset_index(drop=True)
            selected_idx_list = edited_clients.index[edited_clients["_SELECT"] == True].tolist()
            if selected_idx_list:
                local_idx = selected_idx_list[-1]
                base_row = (
                    edited_clients.drop(columns=["_ORIG_INDEX_", "_SELECT"])
                    .iloc[local_idx]
                    .to_dict()
                )
                orig_idx = int(edited_clients["_ORIG_INDEX_"].iloc[local_idx])

                st.markdown(f"**Navette sélectionnée (ligne {local_idx + 1}) :**")
                st.dataframe(
                    edited_clients.drop(columns=["_ORIG_INDEX_", "_SELECT"]).iloc[[local_idx]],
                    use_container_width=True,
                )

    if orig_idx is not None:
        current_id_for_keys = f"row_{orig_idx}"
    else:
        current_id_for_keys = "new"

    st.markdown("### 📝 Fiche de création / modification de navette")

    new_row = {}
    cols_left, cols_right = st.columns(2)

    df2 = st.session_state.df_feuil2
    ch_options = None
    if df2 is not None and not df2.empty and "CH" in df2.columns:
        ch_options = sorted(set(str(x).strip() for x in df2["CH"].dropna() if str(x).strip()))
    elif "CH" in df.columns:
        ch_options = sorted(set(str(x).strip() for x in df["CH"].dropna() if str(x).strip()))

    columns = list(df.columns)

    for i, col in enumerate(columns):
        container = cols_left if i % 2 == 0 else cols_right

        val = base_row.get(col, "") if base_row is not None else ""

        widget_key_prefix = f"client_form_{col}_{current_id_for_keys}"

        if col == "DATE":
            default_date = date.today()
            if isinstance(val, (date, datetime)):
                default_date = val if isinstance(val, date) else val.date()
            else:
                try:
                    ts = pd.to_datetime(val, errors="coerce")
                    if not pd.isna(ts):
                        default_date = ts.date()
                except Exception:
                    default_date = date.today()
            new_val = container.date_input("DATE", value=default_date, key=widget_key_prefix)
            new_row[col] = new_val
            continue

        if col in ["GROUPAGE", "PARTAGE"]:
            bool_default = False
            if isinstance(val, bool):
                bool_default = val
            else:
                try:
                    bool_default = str(val).lower() in ["true", "1", "oui", "x"]
                except Exception:
                    bool_default = False
            label = "Groupage" if col == "GROUPAGE" else "Navette partagée"
            new_val = container.checkbox(
                label,
                value=bool_default,
                key=widget_key_prefix,
            )
            new_row[col] = new_val
            continue

        if col == "CH" and ch_options:
            current = str(val).strip() if val is not None else ""
            opts = [""] + ch_options
            try:
                default_index = opts.index(current)
            except ValueError:
                default_index = 0
            new_val = container.selectbox(
                "CH (chauffeur)",
                opts,
                index=default_index,
                key=widget_key_prefix,
            )
            new_row[col] = new_val
            continue

        if col == "HEURE":
            text_val = "" if val is None or pd.isna(val) else str(val)
            raw_val = container.text_input(col, value=text_val, key=widget_key_prefix)
            new_row[col] = normalize_time_string(raw_val)
            continue

        text_val = "" if val is None or pd.isna(val) else str(val)
        if base_row is None and col == client_col and query.strip():
            text_val = query.strip()
        new_val = container.text_input(col, value=text_val, key=widget_key_prefix)
        new_row[col] = new_val

    groupage_checked = bool(new_row.get("GROUPAGE", False))
    if groupage_checked and client_col:
        st.markdown("### 👥 Groupage — choix du client (onglet Clients)")
        mode = st.radio(
            "Client pour cette navette groupée",
            ["Client existant", "Nouveau client"],
            key=f"groupage_mode_clients_{current_id_for_keys}",
            horizontal=True,
        )
        all_clients = sorted(
            set(str(x).strip() for x in df[client_col].dropna() if str(x).strip())
        )
        if mode == "Client existant":
            selected_client = st.selectbox(
                "Choisir un client existant",
                options=[""] + all_clients,
                key=f"groupage_existing_client_clients_{current_id_for_keys}",
            )
            if selected_client:
                new_row[client_col] = selected_client
        else:
            new_client = st.text_input(
                "Nom du nouveau client pour ce groupage",
                value=str(new_row.get(client_col, "")) if new_row.get(client_col) else "",
                key=f"groupage_new_client_clients_{current_id_for_keys}",
            )
            if new_client:
                new_row[client_col] = new_client

    st.markdown("#### 🚦 Actions sur cette fiche")

    c1, c2, c3 = st.columns(3)

    with c1:
        if st.button("➕ Créer une nouvelle navette pour ce client"):
            df_all = st.session_state.df_feuil1.copy()
            if detect_chauffeur_conflict(df_all, new_row, current_idx=None):
                st.warning("⚠ Conflit chauffeur sur DATE / HEURE, navette créée quand même.")
            df_all = pd.concat([df_all, pd.DataFrame([new_row])], ignore_index=True)
            st.session_state.df_feuil1 = df_all.reset_index(drop=True)
            st.success("Nouvelle navette créée et ajoutée au planning.")
            st.rerun()

    with c2:
        if st.button("✅ Mettre à jour la navette sélectionnée"):
            if base_row is None or orig_idx is None:
                st.warning("Aucune navette existante sélectionnée à mettre à jour.")
            else:
                df_all = st.session_state.df_feuil1.copy()
                if detect_chauffeur_conflict(df_all, new_row, current_idx=orig_idx):
                    st.warning("⚠ Conflit chauffeur sur DATE / HEURE, mise à jour quand même.")
                for k, v in new_row.items():
                    df_all.at[orig_idx, k] = v
                st.session_state.df_feuil1 = df_all.reset_index(drop=True)
                st.success("Navette mise à jour dans le planning.")
                st.rerun()

    with c3:
        if st.button("🗑️ Supprimer la navette sélectionnée"):
            if base_row is None or orig_idx is None:
                st.warning("Aucune navette existante sélectionnée à supprimer.")
            else:
                df_all = st.session_state.df_feuil1.copy()
                df_all = df_all.drop(index=orig_idx).reset_index(drop=True)
                st.session_state.df_feuil1 = df_all
                st.success("Navette supprimée du planning.")
                st.rerun()


# =========================
#   VUE CHAUFFEUR
# =========================

def render_tab_vue_chauffeur():
    st.subheader("🚖 Vue Chauffeur (format texte compact)")

    df = st.session_state.df_feuil1
    df2 = st.session_state.df_feuil2

    if df is None or df.empty:
        st.warning("Aucun planning chargé.")
        return

    if "DATE" in df.columns and not pd.api.types.is_datetime64_any_dtype(df["DATE"]):
        try:
            df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce").dt.date
        except Exception:
            pass

    ch_selected = st.selectbox("Choisir un chauffeur", [""] + CH_CODES, key="vue_chauffeur_ch")
    if not ch_selected:
        st.info("Sélectionne un chauffeur pour voir ses navettes.")
        return

    chauffeur_phone = ""
    chauffeur_email = ""
    if df2 is not None and not df2.empty:
        try:
            mask_ch = df2.iloc[:, 0].astype(str).str.strip() == ch_selected
            match = df2[mask_ch]
            if not match.empty:
                if "TEL_CH" in match.columns:
                    chauffeur_phone = str(match.iloc[0]["TEL_CH"])
                elif match.shape[1] >= 3:
                    chauffeur_phone = str(match.iloc[0, 2])
                # Email colonne 8 (index 7)
                if match.shape[1] >= 8:
                    chauffeur_email = str(match.iloc[0, 7] or "")
        except Exception:
            chauffeur_phone = ""
            chauffeur_email = ""

    today = date.today()
    scope = st.radio(
        "Période à afficher et à envoyer",
        ["Uniquement une date", "À partir de demain (inclus)"],
        index=0,
        horizontal=True,
        key="vue_chauffeur_scope",
    )

    if scope == "Uniquement une date":
        day_selected = st.date_input(
            "Date de la vue chauffeur",
            value=today,
            key="vue_chauffeur_date",
        )
        if "DATE" in df.columns:
            mask_date = df["DATE"] == day_selected
            df_period = df[mask_date].copy()
        else:
            df_period = df.copy()

        if df_period.empty:
            st.warning("Aucune navette pour cette date.")
            return

        def row_has_ch(row):
            ch_list = extract_chauffeurs_from_cell(row.get("CH", ""))
            return ch_selected in ch_list

        df_period["__HAS_CH__"] = df_period.apply(row_has_ch, axis=1)
        df_ch = df_period[df_period["__HAS_CH__"] == True].copy()
        df_ch = df_ch.drop(columns=["__HAS_CH__"])

        if df_ch.empty:
            st.warning(f"Aucune navette pour le chauffeur **{ch_selected}** à cette date.")
            return

        if "HEURE" in df_ch.columns:
            df_ch["__HEURE_SORT__"] = df_ch["HEURE"].apply(sort_heure_for_display)
            df_ch = df_ch.sort_values("__HEURE_SORT__").drop(columns=["__HEURE_SORT__"])

        st.info(f"{len(df_ch)} navette(s) pour **{ch_selected}** le {day_selected.strftime('%d/%m/%Y')}.")
        pdf_bytes = create_chauffeur_pdf(df_ch, ch_selected, day_selected)
        message_txt = build_chauffeur_day_message(df_ch, ch_selected, day_selected)
        mail_subject = f"Planning {day_selected.strftime('%d/%m/%Y')} — {ch_selected}"
        mail_body = message_txt

    else:
        from_date = today + timedelta(days=1)

        if "DATE" in df.columns:
            try:
                df_tmp = df.copy()
                if not pd.api.types.is_datetime64_any_dtype(df_tmp["DATE"]):
                    df_tmp["DATE"] = pd.to_datetime(df_tmp["DATE"], errors="coerce").dt.date
            except Exception:
                df_tmp = df.copy()
        else:
            df_tmp = df.copy()

        if "DATE" in df_tmp.columns:
            df_tmp = df_tmp[df_tmp["DATE"] >= from_date].copy()

        if df_tmp.empty:
            st.warning("Aucune navette à partir de demain.")
            return

        def row_has_ch(row):
            ch_list = extract_chauffeurs_from_cell(row.get("CH", ""))
            return ch_selected in ch_list

        df_tmp["__HAS_CH__"] = df_tmp.apply(row_has_ch, axis=1)
        df_ch = df_tmp[df_tmp["__HAS_CH__"] == True].copy()
        df_ch = df_ch.drop(columns=["__HAS_CH__"])

        if df_ch.empty:
            st.warning(f"Aucune navette pour **{ch_selected}** à partir de demain.")
            return

        if "DATE" in df_ch.columns:
            try:
                df_ch["DATE"] = pd.to_datetime(df_ch["DATE"], errors="coerce").dt.date
            except Exception:
                pass
        if "HEURE" in df_ch.columns:
            df_ch["__HEURE_SORT__"] = df_ch["HEURE"].apply(sort_heure_for_display)
            sort_cols = []
            if "DATE" in df_ch.columns:
                sort_cols.append("DATE")
            sort_cols.append("__HEURE_SORT__")
            df_ch = df_ch.sort_values(sort_cols).drop(columns=["__HEURE_SORT__"])

        st.info(f"{len(df_ch)} navette(s) pour **{ch_selected}** à partir du {from_date.strftime('%d/%m/%Y')}.")
        pdf_bytes = create_chauffeur_pdf(df_ch, ch_selected, from_date)
        message_txt = build_chauffeur_future_message(st.session_state.df_feuil1, ch_selected, from_date)
        mail_subject = f"Planning à partir du {from_date.strftime('%d/%m/%Y')} — {ch_selected}"
        mail_body = message_txt

    col_pdf, col_whats1, col_whats2 = st.columns([1, 1, 1])

    with col_pdf:
        st.download_button(
            "📄 Télécharger la feuille chauffeur (PDF)",
            data=pdf_bytes,
            file_name=f"AirportsLines_{ch_selected}.pdf",
            mime="application/pdf",
        )

    whatsapp_link_msg = build_whatsapp_link(chauffeur_phone, message_txt) if chauffeur_phone else "#"

    with col_whats1:
        if chauffeur_phone:
            st.markdown(
                f"[💬 Envoyer le planning (texte WhatsApp)]({whatsapp_link_msg})",
                unsafe_allow_html=False,
            )
        else:
            st.caption("Pas de numéro chauffeur trouvé (colonne téléphone Feuil2).")

    with col_whats2:
        st.caption("Pour message + PDF : envoie le texte WhatsApp puis joins le PDF téléchargé.")

    st.markdown("---")
    st.markdown("### 📧 Ouvrir un nouveau mail au chauffeur")

    if not chauffeur_email:
        chauffeur_email = st.text_input(
            "Adresse e-mail du chauffeur (colonne 8 Feuil2 ou à remplir ici)",
            value="",
            key=f"chauffeur_email_manual_{ch_selected}",
            placeholder="ex: prenom.nom@domaine.com",
        )
    else:
        chauffeur_email = st.text_input(
            "Adresse e-mail du chauffeur",
            value=chauffeur_email,
            key=f"chauffeur_email_from_feuil2_{ch_selected}",
        )

    if chauffeur_email:
        mailto_link = build_mailto_link(chauffeur_email, mail_subject, mail_body)
        st.markdown(
            f"[📧 Ouvrir Outlook / Mail avec ce planning]({mailto_link})",
            unsafe_allow_html=False,
        )
    else:
        st.caption("Renseigne une adresse e-mail pour activer le bouton.")

    st.markdown("---")
    st.markdown("### 📋 Navettes du chauffeur")

    # AFFICHAGE TEXTE COMPACT, en supprimant les lignes vraiment vides
    for _, row in df_ch.iterrows():
        bloc_lines = []

        heure_txt = normalize_time_string(row.get("HEURE", ""))
        designation = str(row.get("DESIGNATION", "") or "")
        route_text = ""
        for cand in ["Unnamed: 8", "DESIGNATION"]:
            if cand in df_ch.columns:
                val = row.get(cand, "")
                if pd.notna(val) and str(val).strip():
                    route_text = str(val).strip()
                    break

        # Détection des lignes d'indisponibilité chauffeur :
        # si une heure de fin est renseignée dans la colonne "²²²²"
        # et qu'il n'y a pas de client / désignation / trajet, on affiche "INDISPO".
        end_indispo = ""
        if "²²²²" in df_ch.columns:
            end_raw = row.get("²²²²", "")
            end_indispo = normalize_time_string(end_raw)

        is_indispo = False
        if end_indispo:
            nom_client_check = str(row.get("NOM", "") or "").strip()
            if not nom_client_check and not designation and not route_text:
                is_indispo = True

        if is_indispo:
            start_txt = heure_txt or "??:??"
            bloc_lines.append(f"⏱ {start_txt} → {end_indispo} | 🚫 Indisponible")
            bloc_lines.append(f"👨‍✈️ {ch_selected}")
            st.markdown("\n".join(bloc_lines))
            st.markdown("")
            st.markdown("---")
            st.markdown("")
            continue

        if route_text and designation and designation not in route_text:
            dest_full = f"{route_text} ({designation})"
        elif route_text:
            dest_full = route_text
        elif designation:
            dest_full = designation
        else:
            dest_full = ""

        header = f"⏱ {heure_txt or '??:??'}"
        if dest_full:
            header += f" | ✈️ {dest_full}"
        bloc_lines.append(header)

        nom_client = str(row.get("NOM", "") or "").strip()
        if nom_client:
            bloc_lines.append(f"👤 {nom_client}")

        adresse = str(row.get("ADRESSE", "") or "")
        cp = str(row.get("CP", "") or "")
        localite = str(row.get("Localité", "") or row.get("LOCALITE", "") or "")
        adresse_compl = " ".join(x for x in [adresse, cp, localite] if x.strip())
        if adresse_compl:
            bloc_lines.append(f"📍 {adresse_compl}")

        pax = str(row.get("PAX", "") or "").strip()
        immat = str(row.get("IMMAT", "") or "").strip()
        reh = str(row.get("Reh", "") or "").strip()
        siege = str(row.get("Siège", "") or row.get("SIEGE", "") or "").strip()

        veh_parts = []
        if pax:
            veh_parts.append(f"{pax} pax")
        if immat:
            veh_parts.append(f"Immat : {immat}")
        siege_txt = []
        if reh:
            siege_txt.append("Réhausseur")
        if siege:
            if siege.lower() in ["oui", "x", "1"]:
                siege_txt.append("Siège bébé")
            else:
                siege_txt.append(siege)
        if siege_txt:
            veh_parts.append(" / ".join(siege_txt))
        if veh_parts:
            bloc_lines.append("🚐 " + " — ".join(veh_parts))

        tel_client = str(row.get("Tél", "") or row.get("TEL", "") or "").strip()
        if tel_client:
            bloc_lines.append(f"📞 {tel_client}")

        num_vol = str(row.get("N° Vol", "") or "").strip()
        origine = str(row.get("Origine", "") or "").strip()
        decol = str(row.get("Décollage", "") or "").strip()
        h_south = str(row.get("H South", "") or "").strip()

        vol_parts = []
        if num_vol:
            vol_parts.append(f"Vol {num_vol}")
        if origine:
            vol_parts.append(f"Origine : {origine}")
        if decol:
            vol_parts.append(f"Décollage : {decol}")
        if h_south:
            vol_parts.append(f"H SO : {h_south}")
        if vol_parts:
            bloc_lines.append("📝 " + " | ".join(vol_parts))

        paiement = str(row.get("PAIEMENT", "") or "").strip()
        caisse = str(row.get("Caisse", "") or row.get("CAISSE", "") or "").strip()
        num_bdc = str(row.get("Num BDC", "") or "").strip()
        go = str(row.get("GO", "") or "").strip()

        pay_parts = []
        if paiement:
            pay_parts.append(f"Paiement : {paiement}")
        if caisse:
            pay_parts.append(f"Caisse : {caisse}")
        if num_bdc:
            pay_parts.append(f"BDC : {num_bdc}")
        if go:
            pay_parts.append(f"GO : {go}")
        if pay_parts:
            bloc_lines.append("💶 " + " | ".join(pay_parts))

        groupage_flag = False
        partage_flag = False
        if "GROUPAGE" in df_ch.columns:
            try:
                groupage_flag = str(row.get("GROUPAGE", "")).lower() in ["true", "1", "oui", "x"]
            except Exception:
                groupage_flag = False
        if "PARTAGE" in df_ch.columns:
            try:
                partage_flag = str(row.get("PARTAGE", "")).lower() in ["true", "1", "oui", "x"]
            except Exception:
                partage_flag = False

        if groupage_flag:
            bloc_lines.append("🔶 Groupage")
        elif partage_flag:
            bloc_lines.append("🟨 Navette partagée")

        remarque = str(row.get("REMARQUE", "") or "").strip()
        if remarque:
            bloc_lines.append(f"💬 {remarque}")

        bloc_text = "\n".join(bloc_lines)
        st.markdown(bloc_text)
        st.markdown("")
        st.markdown("---")
        st.markdown("")


# =========================
#   MAIN
# =========================

def main():
    init_session_state()

    if not st.session_state.logged_in:
        login_screen()
        return

    if st.session_state.xlsx_bytes is None or st.session_state.df_feuil1 is None:
        load_data_from_ftp()

    render_top_bar()

    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        [
            "📅 Planning",
            "📊 Tableau Feuil1",
            "👨‍✈️ Feuil2 / 📄 Feuil3",
            "🔍 Clients / Historique",
            "🚖 Vue Chauffeur",
        ]
    )

    with tab1:
        render_tab_planning()

    with tab2:
        render_tab_table()

    with tab3:
        colA, colB = st.columns(2)
        with colA:
            render_tab_feuil2()
        with colB:
            render_tab_feuil3()

    with tab4:
        render_tab_clients()

    with tab5:
        render_tab_vue_chauffeur()


if __name__ == "__main__":
    main()
