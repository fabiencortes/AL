# sync_dropbox_to_db.py
"""
SYNC DROPBOX -> SQLITE (SCRIPT STANDALONE)
- PAS de Streamlit
- PAS d'UI
- PAS de rerun
- Mise à jour DB propre
"""

import os
import sqlite3
from datetime import datetime
from io import BytesIO

import pandas as pd
import requests
from openpyxl import load_workbook

# ======================================================
# 🔐 CONFIG
# ======================================================

DB_PATH = "transfers.db"
DROPBOX_PATH = "/Goldenlines/Planning 2026.xlsx"

DROPBOX_APP_KEY = os.environ["DROPBOX_APP_KEY"]
DROPBOX_APP_SECRET = os.environ["DROPBOX_APP_SECRET"]
DROPBOX_REFRESH_TOKEN = os.environ["DROPBOX_REFRESH_TOKEN"]

# ======================================================
# 🔧 DROPBOX
# ======================================================

def get_dropbox_access_token():
    r = requests.post(
        "https://api.dropbox.com/oauth2/token",
        data={
            "grant_type": "refresh_token",
            "refresh_token": DROPBOX_REFRESH_TOKEN,
            "client_id": DROPBOX_APP_KEY,
            "client_secret": DROPBOX_APP_SECRET,
        },
        timeout=10,
    )
    r.raise_for_status()
    return r.json()["access_token"]


def download_dropbox_excel():
    token = get_dropbox_access_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Dropbox-API-Arg": f'{{"path": "{DROPBOX_PATH}"}}',
        "Content-Type": "application/octet-stream",
    }
    r = requests.post(
        "https://content.dropboxapi.com/2/files/download",
        headers=headers,
        timeout=30,
    )
    r.raise_for_status()
    return r.content


# ======================================================
# 🗄️ SQLITE
# ======================================================

def get_connection():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    return conn


# ======================================================
# 🔁 SYNCHRO
# ======================================================

def sync_planning_from_dropbox():
    print("▶️ SYNCHRO START", datetime.now().strftime("%H:%M:%S"))

    # 1️⃣ Télécharger Excel
    content = download_dropbox_excel()
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb["Feuil1"]

    headers = [c.value for c in ws[2]]
    rows = []

    for r in ws.iter_rows(min_row=3, values_only=True):
        if not any(r):
            continue
        rows.append(dict(zip(headers, r)))

    df = pd.DataFrame(rows)

    print(f"📊 {len(df)} lignes lues depuis Excel")

    # 2️⃣ Écrire DB (exemple simple)
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS planning_sync (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            heure TEXT,
            ch TEXT,
            raw TEXT,
            updated_at TEXT
        )
    """)

    cur.execute("DELETE FROM planning_sync")

    for _, row in df.iterrows():
        cur.execute(
            """
            INSERT INTO planning_sync (date, heure, ch, raw, updated_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                str(row.get("DATE")),
                str(row.get("HEURE")),
                str(row.get("CH")),
                str(row),
                datetime.now().isoformat(),
            ),
        )

    conn.commit()
    conn.close()

    print("✅ SYNCHRO OK")
    print("▶️ SYNCHRO END")


# ======================================================
# ▶️ MAIN
# ======================================================

if __name__ == "__main__":
    sync_planning_from_dropbox()
